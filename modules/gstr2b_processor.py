import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ----------------------- Utility Functions ----------------------- #
def get_tax_period(ret_period):
    """Convert return period (e.g., '022025') to month name (e.g., 'February')."""
    print(f"[DEBUG] Parsing tax period from ret_period: {ret_period}")
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    if ret_period and len(ret_period) >= 2:
        result = month_map.get(ret_period[:2], "Unknown")
        print(f"[DEBUG] Tax period parsed: {result}")
        return result
    print("[DEBUG] Tax period parsing failed, returning 'Unknown'")
    return "Unknown"


def parse_number(val, float_2dec=False, int_no_dec=False):
    """Parse a value to a number, handling various formats."""
    print(f"[DEBUG] Parsing number: value={val}, float_2dec={float_2dec}, int_no_dec={int_no_dec}")
    try:
        num = float(val)
        if int_no_dec:
            result = int(num)
            print(f"[DEBUG] Parsed as integer: {result}")
            return result
        if float_2dec:
            result = round(num, 2)
            print(f"[DEBUG] Parsed as float with 2 decimals: {result}")
            return result
        print(f"[DEBUG] Parsed as float: {num}")
        return num
    except (ValueError, TypeError) as e:
        print(f"[DEBUG] Number parsing failed, returning 0: {e}")
        return 0


def get_numeric_value(item, key):
    """Safely extract a numeric value from a dictionary."""
    print(f"[DEBUG] Getting numeric value for key: {key}")
    val = item.get(key, 0)
    if isinstance(val, str):
        val = val.strip()
        print(f"[DEBUG] Stripped string value: {val}")
    result = parse_number(val, float_2dec=True)
    print(f"[DEBUG] Numeric value for {key}: {result}")
    return result


# ----------------------- Extraction Functions ----------------------- #
def extract_b2b(data, filing_period):
    """Extract B2B section data from GSTR-2B JSON (ITC Accepted)."""
    print("[DEBUG] Extracting B2B data...")
    b2b_data = data.get("data", {}).get("docdata", {}).get("b2b", [])
    if not isinstance(b2b_data, list):
        b2b_data = []
        print("[DEBUG] b2b_data is not a list, using empty list")

    extracted_data = {"2B-B2B": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for supplier in b2b_data:
        if not isinstance(supplier, dict):
            print("[DEBUG] Skipping non-dict supplier")
            continue

        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        supprd = supplier.get("supprd", "")
        supfildt = supplier.get("supfildt", "")
        tax_period = get_tax_period(supprd)
        print(f"[DEBUG] Processing supplier: ctin={ctin}, supprd={supprd}, tax_period={tax_period}")

        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list):
            invoices = [invoices] if isinstance(invoices, dict) else []
            print(f"[DEBUG] Converted invoices to list: {len(invoices)} items")

        for inv in invoices:
            if not isinstance(inv, dict):
                print("[DEBUG] Skipping non-dict invoice")
                continue

            print(f"[DEBUG] Processing invoice: inum={inv.get('inum', '')}")
            # Common invoice-level fields
            base_row = {
                "GSTIN of supplier": ctin,
                "Trade/legal name": trdnm,
                "Invoice number": inv.get("inum", ""),
                "Invoice type": inv.get("typ", ""),
                "Invoice Date": inv.get("dt", ""),
                "Invoice Value": get_numeric_value(inv, "val"),
                "Place of supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": inv.get("rev", ""),
                "GSTR Period": tax_period,
                "GSTR Filing Date": supfildt,
                "GSTR Filing Period": filing_period,
                "ITC Availability": inv.get("itcavl", ""),
                "Reason": inv.get("rsn", ""),
                "Source": inv.get("srctyp", "")
            }

            # Check for items (Type 2 JSON)
            items = inv.get("items", [])
            if not isinstance(items, list):
                items = []
                print("[DEBUG] items is not a list, using empty list")

            if items:  # Type 2: Multiple items with different tax rates
                print(f"[DEBUG] Processing Type 2 invoice with {len(items)} items")
                for item in items:
                    if not isinstance(item, dict):
                        print("[DEBUG] Skipping non-dict item")
                        continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""),
                        "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2B"].append(row)
                    print(f"[DEBUG] Added Type 2 row to 2B-B2B: {row}")
            else:  # Type 1: No items, single row
                print("[DEBUG] Processing Type 1 invoice")
                row = base_row.copy()
                row.update({
                    "Invoice Part": "",  # Blank for Type 1
                    "Tax Rate": "",  # Blank for Type 1
                    "Total Taxable Value": get_numeric_value(inv, "txval"),
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2B"].append(row)
                print(f"[DEBUG] Added Type 1 row to 2B-B2B: {row}")

    print("[DEBUG] Finished extracting B2B data")
    return extracted_data


def extract_b2ba(data, filing_period):
    """Extract B2BA section data from GSTR-2B JSON."""
    print("[DEBUG] Extracting B2BA data...")
    b2ba_data = data.get("data", {}).get("docdata", {}).get("b2ba", [])
    if not isinstance(b2ba_data, list):
        b2ba_data = []
        print("[DEBUG] b2ba_data is not a list, using empty list")

    extracted_data = {"2B-B2BA": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for supplier in b2ba_data:
        if not isinstance(supplier, dict):
            print("[DEBUG] Skipping non-dict supplier")
            continue

        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        supprd = supplier.get("supprd", "")
        supfildt = supplier.get("supfildt", "")
        tax_period = get_tax_period(supprd)
        print(f"[DEBUG] Processing supplier: ctin={ctin}, supprd={supprd}, tax_period={tax_period}")

        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list):
            invoices = [invoices] if isinstance(invoices, dict) else []
            print(f"[DEBUG] Converted invoices to list: {len(invoices)} items")

        for inv in invoices:
            if not isinstance(inv, dict):
                print("[DEBUG] Skipping non-dict invoice")
                continue

            print(f"[DEBUG] Processing invoice: inum={inv.get('inum', '')}")
            # Common invoice-level fields
            base_row = {
                "GSTIN of supplier": ctin,
                "Trade/legal name": trdnm,
                "Original Invoice number": inv.get("oinum", ""),
                "Original Invoice Date": inv.get("oidt", ""),
                "Invoice number": inv.get("inum", ""),
                "Invoice Date": inv.get("dt", ""),
                "Invoice type": inv.get("typ", ""),
                "Invoice Value": get_numeric_value(inv, "val"),
                "Place of supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": inv.get("rev", ""),
                "GSTR Period": tax_period,
                "GSTR Filing Date": supfildt,
                "GSTR Filing Period": filing_period,
                "ITC Availability": inv.get("itcavl", ""),
                "Reason": inv.get("rsn", "")
            }

            # Check for items (Type 2 JSON)
            items = inv.get("items", [])
            if not isinstance(items, list):
                items = []
                print("[DEBUG] items is not a list, using empty list")

            if items:  # Type 2: Multiple items with different tax rates
                print(f"[DEBUG] Processing Type 2 invoice with {len(items)} items")
                for item in items:
                    if not isinstance(item, dict):
                        print("[DEBUG] Skipping non-dict item")
                        continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""),
                        "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2BA"].append(row)
                    print(f"[DEBUG] Added Type 2 row to 2B-B2BA: {row}")
            else:  # Type 1: No items, single row
                print("[DEBUG] Processing Type 1 invoice")
                row = base_row.copy()
                row.update({
                    "Invoice Part": "",  # Blank for Type 1
                    "Tax Rate": "",  # Blank for Type 1
                    "Total Taxable Value": get_numeric_value(inv, "txval"),
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2BA"].append(row)
                print(f"[DEBUG] Added Type 1 row to 2B-B2BA: {row}")

    print("[DEBUG] Finished extracting B2BA data")
    return extracted_data


def extract_b2ba_cum(data, filing_period):
    """Extract B2BA cumulative section data from GSTR-2B JSON."""
    print("[DEBUG] Extracting B2BA(cum) data...")
    b2ba_cum_data = data.get("data", {}).get("cpsumm", {}).get("b2ba", [])
    if not isinstance(b2ba_cum_data, list):
        b2ba_cum_data = []
        print("[DEBUG] b2ba_cum_data is not a list, using empty list")

    extracted_data = {"2B-B2BA(cum)": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for supplier in b2ba_cum_data:
        if not isinstance(supplier, dict):
            print("[DEBUG] Skipping non-dict supplier")
            continue

        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        supprd = supplier.get("supprd", "")
        supfildt = supplier.get("supfildt", "")
        tax_period = get_tax_period(supprd)
        print(f"[DEBUG] Processing supplier: ctin={ctin}, supprd={supprd}, tax_period={tax_period}")

        row = {
            "GSTIN of supplier": ctin,
            "Trade/legal name": trdnm,
            "Total Documents": parse_number(supplier.get("ttldocs", 0), int_no_dec=True),
            "Total Taxable Value": get_numeric_value(supplier, "txval"),
            "Integrated Tax": get_numeric_value(supplier, "igst"),
            "Central Tax": get_numeric_value(supplier, "cgst"),
            "State/UT Tax": get_numeric_value(supplier, "sgst"),
            "Cess": get_numeric_value(supplier, "cess"),
            "GSTR Period": tax_period,
            "GSTR Filing Date": supfildt,
            "GSTR Filing Period": filing_period
        }
        extracted_data["2B-B2BA(cum)"].append(row)
        print(f"[DEBUG] Added row to 2B-B2BA(cum): {row}")

    print("[DEBUG] Finished extracting B2BA(cum) data")
    return extracted_data


def extract_cdnr(data, filing_period):
    """Extract CDNR section data from GSTR-2B JSON."""
    print("[DEBUG] Extracting CDNR data...")
    cdnr_data = data.get("data", {}).get("docdata", {}).get("cdnr", [])
    if not isinstance(cdnr_data, list):
        cdnr_data = []
        print("[DEBUG] cdnr_data is not a list, using empty list")

    extracted_data = {"2B-CDNR": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for supplier in cdnr_data:
        if not isinstance(supplier, dict):
            print("[DEBUG] Skipping non-dict supplier")
            continue

        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        supprd = supplier.get("supprd", "")
        supfildt = supplier.get("supfildt", "")
        tax_period = get_tax_period(supprd)
        print(f"[DEBUG] Processing supplier: ctin={ctin}, supprd={supprd}, tax_period={tax_period}")

        notes = supplier.get("nt", [])
        if not isinstance(notes, list):
            notes = [notes] if isinstance(notes, dict) else []
            print(f"[DEBUG] Converted notes to list: {len(notes)} items")

        for note in notes:
            if not isinstance(note, dict):
                print("[DEBUG] Skipping non-dict note")
                continue

            print(f"[DEBUG] Processing note: ntnum={note.get('ntnum', '')}")
            # Common note-level fields
            base_row = {
                "GSTIN of supplier": ctin,
                "Trade/legal name": trdnm,
                "Note number": note.get("ntnum", ""),
                "Note type": note.get("typ", ""),
                "Note supply type": note.get("suptyp", ""),
                "Note Date": note.get("dt", ""),
                "Note Value": get_numeric_value(note, "val"),
                "Place of supply": parse_number(note.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": note.get("rev", ""),
                "GSTR Period": tax_period,
                "GSTR Filing Date": supfildt,
                "GSTR Filing Period": filing_period,
                "ITC Availability": note.get("itcavl", ""),
                "Reason": note.get("rsn", "")
            }

            # Check for items (Type 2 JSON, future-proofing)
            items = note.get("items", [])
            if not isinstance(items, list):
                items = []
                print("[DEBUG] items is not a list, using empty list")

            if items:  # Type 2: Multiple items with different tax rates
                print(f"[DEBUG] Processing Type 2 note with {len(items)} items")
                for item in items:
                    if not isinstance(item, dict):
                        print("[DEBUG] Skipping non-dict item")
                        continue
                    row = base_row.copy()
                    row.update({
                        "Note Part": item.get("num", ""),
                        "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-CDNR"].append(row)
                    print(f"[DEBUG] Added Type 2 row to 2B-CDNR: {row}")
            else:  # Type 1: No items, single row
                print("[DEBUG] Processing Type 1 note")
                row = base_row.copy()
                row.update({
                    "Note Part": "",  # Blank for Type 1
                    "Tax Rate": "",  # Blank for Type 1
                    "Total Taxable Value": get_numeric_value(note, "txval"),
                    "Integrated Tax": get_numeric_value(note, "igst"),
                    "Central Tax": get_numeric_value(note, "cgst"),
                    "State/UT Tax": get_numeric_value(note, "sgst"),
                    "Cess": get_numeric_value(note, "cess")
                })
                extracted_data["2B-CDNR"].append(row)
                print(f"[DEBUG] Added Type 1 row to 2B-CDNR: {row}")

    print("[DEBUG] Finished extracting CDNR data")
    return extracted_data


def extract_impg(data, filing_period):
    """Extract IMPG section data from GSTR-2B JSON."""
    print("[DEBUG] Extracting IMPG data...")
    impg_data = data.get("data", {}).get("docdata", {}).get("impg", [])
    if not isinstance(impg_data, list):
        impg_data = []
        print("[DEBUG] impg_data is not a list, using empty list")

    extracted_data = {"2B-IMPG": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for entry in impg_data:
        if not isinstance(entry, dict):
            print("[DEBUG] Skipping non-dict entry")
            continue

        print(f"[DEBUG] Processing IMPG entry: boenum={entry.get('boenum', '')}")
        row = {
            "ICEGATE Reference Date": entry.get("refdt", ""),
            "Port Code": entry.get("portcode", ""),
            "Bill of Entry Number": entry.get("boenum", ""),
            "Bill of Entry Date": entry.get("boedt", ""),
            "Taxable Value": get_numeric_value(entry, "txval"),
            "Integrated Tax": get_numeric_value(entry, "igst"),
            "Cess": get_numeric_value(entry, "cess"),
            "Record Date": entry.get("recdt", ""),
            "GSTR Filing Period": filing_period,
            "Amended (Yes)": entry.get("isamd", "")
        }
        extracted_data["2B-IMPG"].append(row)
        print(f"[DEBUG] Added row to 2B-IMPG: {row}")

    print("[DEBUG] Finished extracting IMPG data")
    return extracted_data


def extract_b2b_itc_rej(data, filing_period):
    """Extract B2B (ITC Rejected) section data from GSTR-2B JSON."""
    print("[DEBUG] Extracting B2B(ITC_Rej) data...")
    b2b_rej_data = data.get("data", {}).get("docRejdata", {}).get("b2b", [])
    if not isinstance(b2b_rej_data, list):
        b2b_rej_data = []
        print("[DEBUG] b2b_rej_data is not a list, using empty list")

    extracted_data = {"2B-B2B(ITC_Rej)": []}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for supplier in b2b_rej_data:
        if not isinstance(supplier, dict):
            print("[DEBUG] Skipping non-dict supplier")
            continue

        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        supprd = supplier.get("supprd", "")
        supfildt = supplier.get("supfildt", "")
        tax_period = get_tax_period(supprd)
        print(f"[DEBUG] Processing supplier: ctin={ctin}, supprd={supprd}, tax_period={tax_period}")

        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list):
            invoices = [invoices] if isinstance(invoices, dict) else []
            print(f"[DEBUG] Converted invoices to list: {len(invoices)} items")

        for inv in invoices:
            if not isinstance(inv, dict):
                print("[DEBUG] Skipping non-dict invoice")
                continue

            print(f"[DEBUG] Processing invoice: inum={inv.get('inum', '')}")
            # Common invoice-level fields
            base_row = {
                "GSTIN of supplier": ctin,
                "Trade/legal name": trdnm,
                "Invoice number": inv.get("inum", ""),
                "Invoice type": inv.get("typ", ""),
                "Invoice Date": inv.get("dt", ""),
                "Invoice Value": get_numeric_value(inv, "val"),
                "Place of supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "GSTR Period": tax_period,
                "GSTR Filing Date": supfildt,
                "GSTR Filing Period": filing_period,
                "Source": inv.get("srctyp", "")
            }

            # Check for items (Type 2 JSON, future-proofing)
            items = inv.get("items", [])
            if not isinstance(items, list):
                items = []
                print("[DEBUG] items is not a list, using empty list")

            if items:  # Type 2: Multiple items with different tax rates
                print(f"[DEBUG] Processing Type 2 invoice with {len(items)} items")
                for item in items:
                    if not isinstance(item, dict):
                        print("[DEBUG] Skipping non-dict item")
                        continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""),
                        "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2B(ITC_Rej)"].append(row)
                    print(f"[DEBUG] Added Type 2 row to 2B-B2B(ITC_Rej): {row}")
            else:  # Type 1: No items, single row
                print("[DEBUG] Processing Type 1 invoice")
                row = base_row.copy()
                row.update({
                    "Invoice Part": "",  # Blank for Type 1
                    "Tax Rate": "",  # Blank for Type 1
                    "Total Taxable Value": get_numeric_value(inv, "txval"),
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2B(ITC_Rej)"].append(row)
                print(f"[DEBUG] Added Type 1 row to 2B-B2B(ITC_Rej): {row}")

    print("[DEBUG] Finished extracting B2B(ITC_Rej) data")
    return extracted_data


# ----------------------- Summary Generation ----------------------- #
def create_summary_sheets(wb, combined_data):
    """Generate summary sheets from detailed data."""
    print("[DEBUG] Creating summary sheets...")
    summary_data = {
        "2B-Summary-B2B_not RC(ITC_Avl)": [],
        "2B-Summary-B2B_RC(ITC_Avl)": [],
        "2B-Summary-B2BA_not RC(ITC_Avl)": [],
        "2B-Summary-B2BA_RC(ITC_Avl)": [],
        "2B-Summary-B2BA_cum(ITC_Avl)": [],
        "2B-Summary-CDNR_DN(ITC_Avl)": [],
        "2B-Summary-CDNR_CN(ITC_Avl)": [],
        "2B-Summary-CDNR_RC(ITC_Avl)": [],
        "2B-Summary-IMPG(ITC_Avl)": [],
        "2B-Summary-IMPGA(ITC_Avl)": [],
        "2B-Summary-B2B(ITC_Rej)": []
    }

    # Financial year order for sorting
    financial_order = ["April", "May", "June", "July", "August", "September",
                       "October", "November", "December", "January", "February", "March"]

    def has_non_zero_tax(row):
        """Check if any tax field is non-zero."""
        print(f"[DEBUG] Checking non-zero tax for row: {row}")
        result = any(row.get(key, 0) != 0 for key in ["Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])
        print(f"[DEBUG] Non-zero tax check result: {result}")
        return result

    # Collect all unique filing periods from combined_data
    all_filing_periods = set()
    for key, rows in combined_data.items():
        for row in rows:
            filing_period = row.get("GSTR Filing Period", "")
            if filing_period:
                all_filing_periods.add(filing_period)
    print(f"[DEBUG] Collected filing periods: {all_filing_periods}")

    # Initialize monthly summary dictionaries with all filing periods
    monthly_summary_b2b_not_rc = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_b2b_rc = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_b2ba_not_rc = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_b2ba_rc = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_b2ba_cum = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_cdnr_dn = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_cdnr_cn = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_cdnr_rc = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_impg = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_impga = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}
    monthly_summary_b2b_rej = {month: {"Month": month, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for month in all_filing_periods}

    # Aggregate B2B ITC Available data
    b2b_data = combined_data.get("2B-B2B", [])
    print(f"[DEBUG] Total 2B-B2B rows: {len(b2b_data)}")
    b2b_not_rc_count = 0
    b2b_rc_count = 0

    for row in b2b_data:
        month = row["GSTR Filing Period"]
        print(f"[DEBUG] Processing B2B row for month: {month}")
        if row["Supply Attract Reverse Charge"] == "N":
            monthly_summary_b2b_not_rc[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_b2b_not_rc[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_b2b_not_rc[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_b2b_not_rc[month]["Cess"] += row["Cess"]
            b2b_not_rc_count += 1
            print(f"[DEBUG] Aggregated B2B non-RC for {month}")
        elif row["Supply Attract Reverse Charge"] == "Y":
            monthly_summary_b2b_rc[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_b2b_rc[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_b2b_rc[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_b2b_rc[month]["Cess"] += row["Cess"]
            b2b_rc_count += 1
            print(f"[DEBUG] Aggregated B2B RC for {month}")

    print(f"[DEBUG] B2B Non-RC rows processed: {b2b_not_rc_count}")
    print(f"[DEBUG] B2B RC rows processed: {b2b_rc_count}")

    summary_rows_b2b_not_rc = list(monthly_summary_b2b_not_rc.values())
    summary_rows_b2b_not_rc.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2B_not RC(ITC_Avl)"] = summary_rows_b2b_not_rc
    print(f"[DEBUG] B2B Non-RC summary rows: {len(summary_rows_b2b_not_rc)}")

    summary_rows_b2b_rc = list(monthly_summary_b2b_rc.values())
    summary_rows_b2b_rc.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2B_RC(ITC_Avl)"] = summary_rows_b2b_rc
    print(f"[DEBUG] B2B RC summary rows: {len(summary_rows_b2b_rc)}")

    # Aggregate B2BA ITC Available data
    b2ba_data = combined_data.get("2B-B2BA", [])
    print(f"[DEBUG] Total 2B-B2BA rows: {len(b2ba_data)}")
    b2ba_not_rc_count = 0
    b2ba_rc_count = 0

    for row in b2ba_data:
        month = row["GSTR Filing Period"]
        print(f"[DEBUG] Processing B2BA row for month: {month}")
        if row["Supply Attract Reverse Charge"] == "N":
            monthly_summary_b2ba_not_rc[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_b2ba_not_rc[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_b2ba_not_rc[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_b2ba_not_rc[month]["Cess"] += row["Cess"]
            b2ba_not_rc_count += 1
            print(f"[DEBUG] Aggregated B2BA non-RC for {month}")
        elif row["Supply Attract Reverse Charge"] == "Y":
            monthly_summary_b2ba_rc[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_b2ba_rc[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_b2ba_rc[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_b2ba_rc[month]["Cess"] += row["Cess"]
            b2ba_rc_count += 1
            print(f"[DEBUG] Aggregated B2BA RC for {month}")

    print(f"[DEBUG] B2BA Non-RC rows processed: {b2ba_not_rc_count}")
    print(f"[DEBUG] B2BA RC rows processed: {b2ba_rc_count}")

    summary_rows_b2ba_not_rc = list(monthly_summary_b2ba_not_rc.values())
    summary_rows_b2ba_not_rc.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2BA_not RC(ITC_Avl)"] = summary_rows_b2ba_not_rc
    print(f"[DEBUG] B2BA Non-RC summary rows: {len(summary_rows_b2ba_not_rc)}")

    summary_rows_b2ba_rc = list(monthly_summary_b2ba_rc.values())
    summary_rows_b2ba_rc.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2BA_RC(ITC_Avl)"] = summary_rows_b2ba_rc
    print(f"[DEBUG] B2BA RC summary rows: {len(summary_rows_b2ba_rc)}")

    # Aggregate CDNR ITC Available data
    cdnr_data = combined_data.get("2B-CDNR", [])
    print(f"[DEBUG] Total 2B-CDNR rows: {len(cdnr_data)}")
    cdnr_cn_count = 0
    cdnr_dn_count = 0
    cdnr_rc_count = 0

    for row in cdnr_data:
        month = row["GSTR Filing Period"]
        note_type = row["Note type"]
        print(f"[DEBUG] Processing CDNR row for month: {month}, note_type: {note_type}")
        if note_type == "C" and row["Supply Attract Reverse Charge"] != "Y":
            monthly_summary_cdnr_cn[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_cdnr_cn[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_cdnr_cn[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_cdnr_cn[month]["Cess"] += row["Cess"]
            cdnr_cn_count += 1
            print(f"[DEBUG] Aggregated CDNR Credit Note for {month}")
        elif note_type == "D" and row["Supply Attract Reverse Charge"] != "Y":
            monthly_summary_cdnr_dn[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_cdnr_dn[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_cdnr_dn[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_cdnr_dn[month]["Cess"] += row["Cess"]
            cdnr_dn_count += 1
            print(f"[DEBUG] Aggregated CDNR Debit Note for {month}")
        elif row["Supply Attract Reverse Charge"] == "Y":
            monthly_summary_cdnr_rc[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_cdnr_rc[month]["Central Tax"] += row["Central Tax"]
            monthly_summary_cdnr_rc[month]["State/UT Tax"] += row["State/UT Tax"]
            monthly_summary_cdnr_rc[month]["Cess"] += row["Cess"]
            cdnr_rc_count += 1
            print(f"[DEBUG] Aggregated CDNR RC for {month}")

    print(f"[DEBUG] CDNR Credit Note (C) rows processed: {cdnr_cn_count}")
    print(f"[DEBUG] CDNR Debit Note (D) rows processed: {cdnr_dn_count}")
    print(f"[DEBUG] CDNR RC rows processed: {cdnr_rc_count}")

    summary_rows_cdnr_cn = list(monthly_summary_cdnr_cn.values())
    summary_rows_cdnr_cn.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-CDNR_CN(ITC_Avl)"] = summary_rows_cdnr_cn
    print(f"[DEBUG] CDNR Credit Note summary rows: {len(summary_rows_cdnr_cn)}")

    summary_rows_cdnr_dn = list(monthly_summary_cdnr_dn.values())
    summary_rows_cdnr_dn.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-CDNR_DN(ITC_Avl)"] = summary_rows_cdnr_dn
    print(f"[DEBUG] CDNR Debit Note summary rows: {len(summary_rows_cdnr_dn)}")

    summary_rows_cdnr_rc = list(monthly_summary_cdnr_rc.values())
    summary_rows_cdnr_rc.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-CDNR_RC(ITC_Avl)"] = summary_rows_cdnr_rc
    print(f"[DEBUG] CDNR RC summary rows: {len(summary_rows_cdnr_rc)}")

    # Aggregate B2BA_cum ITC Available data
    b2ba_cum_data = combined_data.get("2B-B2BA(cum)", [])
    print(f"[DEBUG] Total 2B-B2BA(cum) rows: {len(b2ba_cum_data)}")
    b2ba_cum_count = 0

    for row in b2ba_cum_data:
        month = row["GSTR Filing Period"]
        print(f"[DEBUG] Processing B2BA(cum) row for month: {month}")
        monthly_summary_b2ba_cum[month]["Integrated Tax"] += row["Integrated Tax"]
        monthly_summary_b2ba_cum[month]["Central Tax"] += row["Central Tax"]
        monthly_summary_b2ba_cum[month]["State/UT Tax"] += row["State/UT Tax"]
        monthly_summary_b2ba_cum[month]["Cess"] += row["Cess"]
        b2ba_cum_count += 1
        print(f"[DEBUG] Aggregated B2BA(cum) for {month}")

    print(f"[DEBUG] B2BA(cum) rows processed: {b2ba_cum_count}")
    summary_rows_b2ba_cum = list(monthly_summary_b2ba_cum.values())
    summary_rows_b2ba_cum.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2BA_cum(ITC_Avl)"] = summary_rows_b2ba_cum
    print(f"[DEBUG] B2BA(cum) summary rows: {len(summary_rows_b2ba_cum)}")

    # Aggregate IMPG ITC Available data (Non-amended)
    impg_data = combined_data.get("2B-IMPG", [])
    print(f"[DEBUG] Total 2B-IMPG rows: {len(impg_data)}")
    impg_not_amd_count = 0
    impga_amd_count = 0

    for row in impg_data:
        month = row["GSTR Filing Period"]
        print(f"[DEBUG] Processing IMPG row for month: {month}")
        if row["Amended (Yes)"] == "N":
            monthly_summary_impg[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_impg[month]["Central Tax"] += 0  # Always 0 for IMPG
            monthly_summary_impg[month]["State/UT Tax"] += 0  # Always 0 for IMPG
            monthly_summary_impg[month]["Cess"] += row["Cess"]
            impg_not_amd_count += 1
            print(f"[DEBUG] Aggregated IMPG non-amended for {month}")
        elif row["Amended (Yes)"] == "Y":
            monthly_summary_impga[month]["Integrated Tax"] += row["Integrated Tax"]
            monthly_summary_impga[month]["Central Tax"] += 0  # Always 0 for IMPG
            monthly_summary_impga[month]["State/UT Tax"] += 0  # Always 0 for IMPG
            monthly_summary_impga[month]["Cess"] += row["Cess"]
            impga_amd_count += 1
            print(f"[DEBUG] Aggregated IMPG amended for {month}")

    print(f"[DEBUG] IMPG Non-amended rows processed: {impg_not_amd_count}")
    print(f"[DEBUG] IMPG Amended rows processed: {impga_amd_count}")

    summary_rows_impg = list(monthly_summary_impg.values())
    summary_rows_impg.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-IMPG(ITC_Avl)"] = summary_rows_impg
    print(f"[DEBUG] IMPG summary rows: {len(summary_rows_impg)}")

    summary_rows_impga = list(monthly_summary_impga.values())
    summary_rows_impga.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-IMPGA(ITC_Avl)"] = summary_rows_impga
    print(f"[DEBUG] IMPGA summary rows: {len(summary_rows_impga)}")

    # Aggregate B2B ITC Rejected data
    b2b_rej_data = combined_data.get("2B-B2B(ITC_Rej)", [])
    print(f"[DEBUG] Total 2B-B2B(ITC_Rej) rows: {len(b2b_rej_data)}")
    b2b_rej_count = 0

    for row in b2b_rej_data:
        month = row["GSTR Filing Period"]
        print(f"[DEBUG] Processing B2B(ITC_Rej) row for month: {month}")
        monthly_summary_b2b_rej[month]["Integrated Tax"] += row["Integrated Tax"]
        monthly_summary_b2b_rej[month]["Central Tax"] += row["Central Tax"]
        monthly_summary_b2b_rej[month]["State/UT Tax"] += row["State/UT Tax"]
        monthly_summary_b2b_rej[month]["Cess"] += row["Cess"]
        b2b_rej_count += 1
        print(f"[DEBUG] Aggregated B2B(ITC_Rej) for {month}")

    print(f"[DEBUG] B2B(ITC_Rej) rows processed: {b2b_rej_count}")
    summary_rows_b2b_rej = list(monthly_summary_b2b_rej.values())
    summary_rows_b2b_rej.sort(key=lambda x: financial_order.index(x["Month"]) if x["Month"] in financial_order else 999)
    summary_data["2B-Summary-B2B(ITC_Rej)"] = summary_rows_b2b_rej
    print(f"[DEBUG] B2B(ITC_Rej) summary rows: {len(summary_rows_b2b_rej)}")

    # Summary sheet definitions
    section_titles = {
        "2B-Summary-B2B_not RC(ITC_Avl)": "3. ITC Available - PART A1 - Supplies other than Reverse charge - B2B Invoices (IMS) - Summary",
        "2B-Summary-B2B_RC(ITC_Avl)": "3. ITC Available - PART A3 - Supplies liable for Reverse charge - B2B Invoices - Summary",
        "2B-Summary-B2BA_not RC(ITC_Avl)": "3. ITC Available - PART A1 - Supplies other than Reverse charge - B2BA Invoices (IMS) - Summary",
        "2B-Summary-B2BA_RC(ITC_Avl)": "3. ITC Available - PART A3 - Supplies liable for Reverse charge - B2BA Invoices (IMS) - Summary",
        "2B-Summary-CDNR_CN(ITC_Avl)": "3. ITC Available - PART B1 - B2B Credit Notes (IMS) - Summary",
        "2B-Summary-CDNR_RC(ITC_Avl)": "3. ITC Available - PART B1 - B2B Credit Notes (Reverse charge) - Summary",
        "2B-Summary-B2BA_cum(ITC_Avl)": "3. ITC Available - PART A1&A3 - All Supplies including those liable for Reverse charge - B2BA(cum) Invoices (IMS) - Summary",
        "2B-Summary-IMPG(ITC_Avl)": "3. ITC Available - PART A4 - Import of goods from overseas - Summary",
        "2B-Summary-IMPGA(ITC_Avl)": "3. ITC Available - PART A4 - Import of goods from overseas (Amendment) - Summary",
        "2B-Summary-B2B(ITC_Rej)": "6. ITC Rejected - PART A1 - Supplies other than Reverse charge - B2B Invoices (IMS) - Summary",
        "2B-Summary-CDNR_DN(ITC_Avl)": "3. ITC Available - PART A1 - B2B Debit Notes - Summary"
    }
    column_headers = {
        "2B-Summary-B2B_not RC(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-B2B_RC(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-B2BA_not RC(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-B2BA_RC(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-CDNR_CN(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-CDNR_RC(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-B2BA_cum(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-IMPG(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-IMPGA(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-B2B(ITC_Rej)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "2B-Summary-CDNR_DN(ITC_Avl)": ["Month", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"]
    }
    indian_format = r"[>=10000000]##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    column_formats = {
        "2B-Summary-B2B_not RC(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-B2B_RC(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-B2BA_not RC(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-B2BA_RC(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-CDNR_CN(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-CDNR_RC(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-B2BA_cum(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-IMPG(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-IMPGA(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-B2B(ITC_Rej)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        },
        "2B-Summary-CDNR_DN(ITC_Avl)": {
            "Month": "General",
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format
        }
    }

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    def sheet_has_valid_data(rows, numeric_headers):
        """Check if the sheet has non-zero numeric data."""
        print(f"[DEBUG] Checking if sheet has valid data, numeric headers: {numeric_headers}")
        for row in rows:
            for header in numeric_headers:
                try:
                    if float(row.get(header, 0)) != 0:
                        print(f"[DEBUG] Found non-zero value for {header}: {row.get(header)}")
                        return True
                except (ValueError, TypeError):
                    continue
        print("[DEBUG] No valid data found for sheet")
        return False

    for sheet_key, rows in summary_data.items():
        numeric_headers = ["Integrated Tax", "Central Tax", "State/UT Tax", "Cess"]
        print(f"[DEBUG] Processing summary sheet: {sheet_key}, rows: {len(rows)}")
        if not sheet_has_valid_data(rows, numeric_headers):
            print(f"[DEBUG] Skipping {sheet_key} due to no valid data")
            continue

        print(f"[DEBUG] Creating sheet: {sheet_key} with {len(rows)} rows")
        if sheet_key in wb.sheetnames:
            print(f"[DEBUG] Removing existing sheet: {sheet_key}")
            wb.remove(wb[sheet_key])

        ws = wb.create_sheet(sheet_key)
        print(f"[DEBUG] Created sheet: {sheet_key}")
        ws.freeze_panes = "B3"  # Freeze at B3 (locks rows 1-2 and column A)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_headers[sheet_key]))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = section_titles[sheet_key]
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        print(f"[DEBUG] Set sheet title: {section_titles[sheet_key]}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        print(f"[DEBUG] Added headers: {column_headers[sheet_key]}")

        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
                cell_value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if col_name in column_formats[sheet_key]:
                    cell.number_format = column_formats[sheet_key][col_name]
        print(f"[DEBUG] Populated data rows for {sheet_key}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(col_name)
            for row in rows:
                max_length = max(max_length, len(str(row.get(col_name, ""))))
            ws.column_dimensions[col_letter].width = max(15, max_length + 2)
        print(f"[DEBUG] Adjusted column widths for {sheet_key}")

    print("[DEBUG] Finished creating summary sheets")


# ----------------------- Excel Report Generation ----------------------- #
def create_excel_report(data_dict, save_path, template_path=None):
    """Generate an Excel report from extracted GSTR-2B data with freeze panes."""
    print("[DEBUG] Starting Excel report creation...")
    if template_path and os.path.exists(template_path):
        print(f"[DEBUG] Loading template from: {template_path}")
        wb = load_workbook(template_path)
    else:
        print("[DEBUG] Creating new workbook")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            print("[DEBUG] Removed default 'Sheet'")
            wb.remove(wb["Sheet"])

    section_titles = {
        "2B-B2B": "Taxable inward supplies received from registered person : 2B-B2B",
        "2B-B2BA": "Amendments to previously filed invoices by supplier : 2B-B2BA",
        "2B-B2BA(cum)": "Amendments to previously filed invoices by supplier : 2B-B2BA (Cumulative)",
        "2B-CDNR": "Debit/Credit notes(Original) : 2B-CDNR",
        "2B-IMPG": "Import of Goods : 2B-IMPG",
        "2B-B2B(ITC_Rej)": "Taxable inward supplies received from registered person : 2B-B2B (ITC Rejected)"
    }

    column_headers = {
        "2B-B2B": [
            "GSTIN of supplier", "Trade/legal name", "Invoice number", "Invoice Part",
            "Invoice type", "Tax Rate", "Invoice Date", "Invoice Value", "Place of supply",
            "Supply Attract Reverse Charge", "Total Taxable Value", "Integrated Tax",
            "Central Tax", "State/UT Tax", "Cess", "GSTR Period", "GSTR Filing Date",
            "GSTR Filing Period", "ITC Availability", "Reason", "Source"
        ],
        "2B-B2BA": [
            "GSTIN of supplier", "Trade/legal name", "Original Invoice number", "Original Invoice Date",
            "Invoice number", "Invoice Part", "Invoice Date", "Invoice type", "Tax Rate",
            "Invoice Value", "Place of supply", "Supply Attract Reverse Charge", "Total Taxable Value",
            "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "GSTR Period", "GSTR Filing Date",
            "GSTR Filing Period", "ITC Availability", "Reason"
        ],
        "2B-B2BA(cum)": [
            "GSTIN of supplier", "Trade/legal name", "Total Documents", "Total Taxable Value",
            "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "GSTR Period",
            "GSTR Filing Date", "GSTR Filing Period"
        ],
        "2B-CDNR": [
            "GSTIN of supplier", "Trade/legal name", "Note number", "Note Part", "Note type",
            "Tax Rate", "Note supply type", "Note Date", "Note Value", "Place of supply",
            "Supply Attract Reverse Charge", "Total Taxable Value", "Integrated Tax",
            "Central Tax", "State/UT Tax", "Cess", "GSTR Period", "GSTR Filing Date",
            "GSTR Filing Period", "ITC Availability", "Reason"
        ],
        "2B-IMPG": [
            "ICEGATE Reference Date", "Port Code", "Bill of Entry Number", "Bill of Entry Date",
            "Taxable Value", "Integrated Tax", "Cess", "Record Date", "GSTR Filing Period",
            "Amended (Yes)"
        ],
        "2B-B2B(ITC_Rej)": [
            "GSTIN of supplier", "Trade/legal name", "Invoice number", "Invoice Part",
            "Invoice type", "Tax Rate", "Invoice Date", "Invoice Value", "Place of supply",
            "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
            "GSTR Period", "GSTR Filing Date", "GSTR Filing Period", "Source"
        ]
    }

    indian_format = r"[>=10000000]##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    column_formats = {
        "2B-B2B": {
            "GSTIN of supplier": "General",
            "Trade/legal name": "General",
            "Invoice number": "General",
            "Invoice Part": "General",
            "Invoice type": "General",
            "Tax Rate": "General",
            "Invoice Date": "dd-mm-yyyy",
            "Invoice Value": indian_format,
            "Place of supply": "General",
            "Supply Attract Reverse Charge": "General",
            "Total Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format,
            "GSTR Period": "General",
            "GSTR Filing Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General",
            "ITC Availability": "General",
            "Reason": "General",
            "Source": "General"
        },
        "2B-B2BA": {
            "GSTIN of supplier": "General",
            "Trade/legal name": "General",
            "Original Invoice number": "General",
            "Original Invoice Date": "dd-mm-yyyy",
            "Invoice number": "General",
            "Invoice Part": "General",
            "Invoice Date": "dd-mm-yyyy",
            "Invoice type": "General",
            "Tax Rate": "General",
            "Invoice Value": indian_format,
            "Place of supply": "General",
            "Supply Attract Reverse Charge": "General",
            "Total Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format,
            "GSTR Period": "General",
            "GSTR Filing Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General",
            "ITC Availability": "General",
            "Reason": "General"
        },
        "2B-B2BA(cum)": {
            "GSTIN of supplier": "General",
            "Trade/legal name": "General",
            "Total Documents": "0",
            "Total Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format,
            "GSTR Period": "General",
            "GSTR Filing Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General"
        },
        "2B-CDNR": {
            "GSTIN of supplier": "General",
            "Trade/legal name": "General",
            "Note number": "General",
            "Note Part": "General",
            "Note type": "General",
            "Tax Rate": "General",
            "Note supply type": "General",
            "Note Date": "dd-mm-yyyy",
            "Note Value": indian_format,
            "Place of supply": "General",
            "Supply Attract Reverse Charge": "General",
            "Total Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format,
            "GSTR Period": "General",
            "GSTR Filing Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General",
            "ITC Availability": "General",
            "Reason": "General"
        },
        "2B-IMPG": {
            "ICEGATE Reference Date": "dd-mm-yyyy",
            "Port Code": "General",
            "Bill of Entry Number": "General",
            "Bill of Entry Date": "dd-mm-yyyy",
            "Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Cess": indian_format,
            "Record Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General",
            "Amended (Yes)": "General"
        },
        "2B-B2B(ITC_Rej)": {
            "GSTIN of supplier": "General",
            "Trade/legal name": "General",
            "Invoice number": "General",
            "Invoice Part": "General",
            "Invoice type": "General",
            "Tax Rate": "General",
            "Invoice Date": "dd-mm-yyyy",
            "Invoice Value": indian_format,
            "Place of supply": "General",
            "Total Taxable Value": indian_format,
            "Integrated Tax": indian_format,
            "Central Tax": indian_format,
            "State/UT Tax": indian_format,
            "Cess": indian_format,
            "GSTR Period": "General",
            "GSTR Filing Date": "dd-mm-yyyy",
            "GSTR Filing Period": "General",
            "Source": "General"
        }
    }

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    def sheet_has_valid_data(rows, numeric_headers):
        """Check if the sheet has non-zero numeric data."""
        print(f"[DEBUG] Checking if sheet has valid data, numeric headers: {numeric_headers}")
        for row in rows:
            for header in numeric_headers:
                try:
                    if float(row.get(header, 0)) != 0:
                        print(f"[DEBUG] Found non-zero value for {header}: {row.get(header)}")
                        return True
                except (ValueError, TypeError):
                    continue
        print("[DEBUG] No valid data found for sheet")
        return False

    print("[DEBUG] Generating sheets...")
    for sheet_key, rows in data_dict.items():
        numeric_headers = [
            "Total Taxable Value" if sheet_key in ["2B-B2BA(cum)", "2B-IMPG"] else
            "Note Value" if sheet_key == "2B-CDNR" else "Invoice Value",
            "Total Taxable Value" if sheet_key != "2B-IMPG" else "Taxable Value",
            "Integrated Tax", "Central Tax" if sheet_key != "2B-IMPG" else None,
            "State/UT Tax" if sheet_key != "2B-IMPG" else None, "Cess"
        ]
        numeric_headers = [h for h in numeric_headers if h]  # Remove None
        print(f"[DEBUG] Processing sheet: {sheet_key}, rows: {len(rows)}")
        if not sheet_has_valid_data(rows, numeric_headers):
            print(f"[DEBUG] Skipping sheet {sheet_key} due to no valid data")
            continue

        if sheet_key in wb.sheetnames:
            print(f"[DEBUG] Removing existing sheet: {sheet_key}")
            wb.remove(wb[sheet_key])

        ws = wb.create_sheet(sheet_key)
        print(f"[DEBUG] Created sheet: {sheet_key}")
        ws.freeze_panes = "B3"  # Freeze at B3 (locks rows 1-2 and column A)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_headers[sheet_key]))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = section_titles[sheet_key]
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        print(f"[DEBUG] Set sheet title: {section_titles[sheet_key]}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        print(f"[DEBUG] Added headers: {column_headers[sheet_key]}")

        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
                cell_value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if col_name in column_formats[sheet_key]:
                    cell.number_format = column_formats[sheet_key][col_name]
        print(f"[DEBUG] Populated data rows for {sheet_key}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(col_name)
            for row in rows:
                max_length = max(max_length, len(str(row.get(col_name, ""))))
            ws.column_dimensions[col_letter].width = max(15, max_length + 2)
        print(f"[DEBUG] Adjusted column widths for {sheet_key}")

    # Add summary sheets
    create_summary_sheets(wb, data_dict)

    print(f"[DEBUG] Saving workbook to {save_path}...")
    wb.save(save_path)
    print(f"[DEBUG] Workbook saved successfully to {save_path}")
    print("[DEBUG] Finished Excel report creation")
    return f"✅ Successfully saved Excel file: {save_path}"


# ----------------------- Main Processing Function ----------------------- #
def process_gstr2b(json_files, template_path, save_path):
    """Process GSTR-2B JSON files and generate Excel report."""
    print("[DEBUG] Starting GSTR-2B processing...")
    combined_data = {
        "2B-B2B": [],
        "2B-B2BA": [],
        "2B-B2BA(cum)": [],
        "2B-CDNR": [],
        "2B-IMPG": [],
        "2B-B2B(ITC_Rej)": []
    }
    print(f"[DEBUG] Initialized combined_data with keys: {list(combined_data.keys())}")

    for file_path in json_files:
        print(f"[DEBUG] Processing JSON file: {file_path}")
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            filing_period = get_tax_period(data.get("data", {}).get("rtnprd", ""))
            print(f"[DEBUG] Loaded file {file_path}, Filing Period: {filing_period}")
            extracted_b2b = extract_b2b(data, filing_period)
            extracted_b2ba = extract_b2ba(data, filing_period)
            extracted_b2ba_cum = extract_b2ba_cum(data, filing_period)
            extracted_cdnr = extract_cdnr(data, filing_period)
            extracted_impg = extract_impg(data, filing_period)
            extracted_b2b_rej = extract_b2b_itc_rej(data, filing_period)
            for key, rows in extracted_b2b.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")
            for key, rows in extracted_b2ba.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")
            for key, rows in extracted_b2ba_cum.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")
            for key, rows in extracted_cdnr.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")
            for key, rows in extracted_impg.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")
            for key, rows in extracted_b2b_rej.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

    print("[DEBUG] Sorting data...")
    financial_order = ["April", "May", "June", "July", "August", "September",
                       "October", "November", "December", "January", "February", "March"]
    for key in combined_data:
        if key in ["2B-B2B", "2B-B2BA", "2B-B2BA(cum)", "2B-CDNR", "2B-B2B(ITC_Rej)"]:
            print(f"[DEBUG] Sorting section: {key}")
            combined_data[key].sort(
                key=lambda x: financial_order.index(x["GSTR Period"])
                if x["GSTR Period"] in financial_order else 999
            )
            print(f"[DEBUG] Sorted {key} with {len(combined_data[key])} rows")

    print("[DEBUG] Generating Excel report...")
    result = create_excel_report(combined_data, save_path, template_path)
    print("[DEBUG] GSTR-2B processing completed")
    return result