import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ----------------------- Utility Functions ----------------------- #
def get_tax_period(ret_period):
    print(f"[DEBUG] Parsing tax period from ret_period: {ret_period}")
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    if ret_period and len(ret_period) >= 2:
        result = month_map.get(ret_period[:2], "Unknown")
        print(f"[DEBUG] Tax period parsed: {result}")
        return result
    print("[DEBUG] Tax period parsing failed, returning 'Unknown'")
    return "Unknown"

def get_tax_period_from_date(trandate):
    print(f"[DEBUG] Parsing tax period from date: {trandate}")
    try:
        parts = trandate.split("-")
        if len(parts) == 3:
            month = parts[1]
            month_map = {
                "01": "January", "02": "February", "03": "March", "04": "April",
                "05": "May", "06": "June",
                "07": "July", "08": "August", "09": "September",
                "10": "October", "11": "November", "12": "December"
            }
            result = month_map.get(month, "Unknown")
            print(f"[DEBUG] Tax period from date parsed: {result}")
            return result
    except Exception as e:
        print(f"[DEBUG] Error parsing tax period from date: {e}")
    print("[DEBUG] Tax period from date parsing failed, returning 'Unknown'")
    return "Unknown"

def parse_number(val, float_2dec=False, int_no_dec=False):
    print(f"[DEBUG] Parsing number: value={val}, float_2dec={float_2dec}, int_no_dec={int_no_dec}")
    try:
        num = float(val)
        if int_no_dec:
            result = int(num)
            print(f"[DEBUG] Parsed as integer: {result}")
            return result
        if float_2dec:
            result = round(num, 2)
            print(f"[DEBUG] Parsed as float with 2 decimals: {result}")
            return result
        print(f"[DEBUG] Parsed as float: {num}")
        return num
    except (ValueError, TypeError) as e:
        print(f"[DEBUG] Number parsing failed, returning 0: {e}")
        return 0

def get_numeric_value(item, key):
    print(f"[DEBUG] Getting numeric value for key: {key}")
    val = item.get(key, 0)
    if isinstance(val, str):
        val = val.strip()
        print(f"[DEBUG] Stripped string value: {val}")
    result = parse_number(val, float_2dec=True)
    print(f"[DEBUG] Numeric value for {key}: {result}")
    return result

# ----------------------- Extraction Functions ----------------------- #
def extract_section_3_1(data):
    print("[DEBUG] Extracting section 3.1 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    sup_details = r3b_data.get("sup_details", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    section_mapping = {
        "OSUP-Detail": "osup_det",
        "OSUP-Zero": "osup_zero",
        "OSUP-Nil,Exmp": "osup_nil_exmp",
        "ISUP-Rev": "isup_rev",
        "OSUP-NonGST": "osup_nongst",
    }

    extracted_data = {key: [] for key in section_mapping}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for sheet_name, json_key in section_mapping.items():
        print(f"[DEBUG] Processing sheet: {sheet_name}, JSON key: {json_key}")
        section_data = sup_details.get(json_key, {})
        if not isinstance(section_data, dict):
            section_data = {}
            print(f"[DEBUG] Section data for {json_key} is not a dict, using empty dict")
        row = {
            "Tax Period": tax_period,
            "Total taxable value": get_numeric_value(section_data, "txval"),
            "Integrated Tax": get_numeric_value(section_data, "iamt"),
            "Central Tax": get_numeric_value(section_data, "camt"),
            "State/UT Tax": get_numeric_value(section_data, "samt"),
            "Cess": get_numeric_value(section_data, "csamt")
        }
        extracted_data[sheet_name].append(row)
        print(f"[DEBUG] Added row to {sheet_name}: {row}")

    # Add default rows for empty sheets
    for sheet_name in section_mapping:
        if not extracted_data[sheet_name]:
            print(f"[DEBUG] No data for {sheet_name} in tax period {tax_period}, adding default row")
            default_row = {
                "Tax Period": tax_period,
                "Total taxable value": 0.00,
                "Integrated Tax": 0.00,
                "Central Tax": 0.00,
                "State/UT Tax": 0.00,
                "Cess": 0.00
            }
            extracted_data[sheet_name].append(default_row)
            print(f"[DEBUG] Added default row to {sheet_name}: {default_row}")

    print("[DEBUG] Finished extracting section 3.1 data")
    return extracted_data

def extract_section_3_2(data):
    print("[DEBUG] Extracting section 3.2 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    inter_sup = r3b_data.get("inter_sup", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    section_mapping = {
        "InterSUP-Unreg": "unreg_details",
        "InterSUP-Comp": "comp_details",
        "InterSUP-UIN": "uin_details",
    }

    extracted_data = {key: [] for key in section_mapping}
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    for sheet_name, json_key in section_mapping.items():
        print(f"[DEBUG] Processing sheet: {sheet_name}, JSON key: {json_key}")
        section_data = inter_sup.get(json_key, [])
        if not isinstance(section_data, list):
            section_data = [section_data] if isinstance(section_data, dict) else []
            print(f"[DEBUG] Converted {json_key} data to list: {len(section_data)} items")
        for item in section_data:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping non-dict item in {sheet_name}")
                continue
            row = {
                "Tax Period": tax_period,
                "Total taxable value": get_numeric_value(item, "txval"),
                "Integrated Tax": get_numeric_value(item, "iamt")
            }
            extracted_data[sheet_name].append(row)
            print(f"[DEBUG] Added row to {sheet_name}: {row}")

    # Add default rows for empty sheets
    for sheet_name in section_mapping:
        if not extracted_data[sheet_name]:
            print(f"[DEBUG] No data for {sheet_name} in tax period {tax_period}, adding default row")
            default_row = {
                "Tax Period": tax_period,
                "Total taxable value": 0.00,
                "Integrated Tax": 0.00
            }
            extracted_data[sheet_name].append(default_row)
            print(f"[DEBUG] Added default row to {sheet_name}: {default_row}")

    print("[DEBUG] Finished extracting section 3.2 data")
    return extracted_data

def extract_section_4(data):
    print("[DEBUG] Extracting section 4 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    itc_elg = r3b_data.get("itc_elg", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    extracted_data = {
        "ITC-Available": [],
        "ITC-avl-IMPG": [],
        "ITC-avl-IMPS": [],
        "ITC-avl-ISRC": [],
        "ITC-avl-ISD": [],
        "ITC-avl-OTH": [],
        "ITC-Reversed": [],
        "ITC-rev-RUL": [],
        "ITC-rev-OTH": [],
        "Net-ITC": [],
        "ITC-Ineligible": [],
        "ITC-inelg-RUL": [],
        "ITC-inelg-OTH": [],
    }
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    itc_avl = itc_elg.get("itc_avl", [])
    if not isinstance(itc_avl, list):
        itc_avl = [itc_avl] if isinstance(itc_avl, dict) else []
        print(f"[DEBUG] Converted itc_avl to list: {len(itc_avl)} items")
    summary_avl = {}
    for item in itc_avl:
        if not isinstance(item, dict):
            print("[DEBUG] Skipping non-dict item in itc_avl")
            continue
        sub_type = item.get("ty", "")
        if isinstance(sub_type, str):
            sub_type = sub_type.strip()
            print(f"[DEBUG] Processing itc_avl item with type: {sub_type}")
        row = {
            "Tax Period": tax_period,
            "Integrated Tax": get_numeric_value(item, "iamt"),
            "Central Tax": get_numeric_value(item, "camt"),
            "State/UT Tax": get_numeric_value(item, "samt"),
            "Cess": get_numeric_value(item, "csamt")
        }
        key = f"ITC-avl-{sub_type}"
        if key in extracted_data:
            extracted_data[key].append(row)
            print(f"[DEBUG] Added row to {key}: {row}")
        if tax_period not in summary_avl:
            summary_avl[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_avl[tax_period]["Integrated Tax"] += row["Integrated Tax"]
        summary_avl[tax_period]["Central Tax"] += row["Central Tax"]
        summary_avl[tax_period]["State/UT Tax"] += row["State/UT Tax"]
        summary_avl[tax_period]["Cess"] += row["Cess"]

    for sp in summary_avl.values():
        extracted_data["ITC-Available"].append(sp)
        print(f"[DEBUG] Added summary row to ITC-Available: {sp}")

    itc_rev = itc_elg.get("itc_rev", [])
    if not isinstance(itc_rev, list):
        itc_rev = [itc_rev] if isinstance(itc_rev, dict) else []
        print(f"[DEBUG] Converted itc_rev to list: {len(itc_rev)} items")
    summary_rev = {}
    for item in itc_rev:
        if not isinstance(item, dict):
            print("[DEBUG] Skipping non-dict item in itc_rev")
            continue
        sub_type = item.get("ty", "")
        if isinstance(sub_type, str):
            sub_type = sub_type.strip()
            print(f"[DEBUG] Processing itc_rev item with type: {sub_type}")
        row = {
            "Tax Period": tax_period,
            "Integrated Tax": get_numeric_value(item, "iamt"),
            "Central Tax": get_numeric_value(item, "camt"),
            "State/UT Tax": get_numeric_value(item, "samt"),
            "Cess": get_numeric_value(item, "csamt")
        }
        key = f"ITC-rev-{sub_type}"
        if key in extracted_data:
            extracted_data[key].append(row)
            print(f"[DEBUG] Added row to {key}: {row}")
        if tax_period not in summary_rev:
            summary_rev[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_rev[tax_period]["Integrated Tax"] += row["Integrated Tax"]
        summary_rev[tax_period]["Central Tax"] += row["Central Tax"]
        summary_rev[tax_period]["State/UT Tax"] += row["State/UT Tax"]
        summary_rev[tax_period]["Cess"] += row["Cess"]

    for sp in summary_rev.values():
        extracted_data["ITC-Reversed"].append(sp)
        print(f"[DEBUG] Added summary row to ITC-Reversed: {sp}")

    itc_net = itc_elg.get("itc_net", {})
    if not isinstance(itc_net, dict):
        itc_net = {}
        print("[DEBUG] itc_net is not a dict, using empty dict")
    row = {
        "Tax Period": tax_period,
        "Integrated Tax": get_numeric_value(itc_net, "iamt"),
        "Central Tax": get_numeric_value(itc_net, "camt"),
        "State/UT Tax": get_numeric_value(itc_net, "samt"),
        "Cess": get_numeric_value(itc_net, "csamt")
    }
    extracted_data["Net-ITC"].append(row)
    print(f"[DEBUG] Added row to Net-ITC: {row}")

    itc_inelg = itc_elg.get("itc_inelg", [])
    if not isinstance(itc_inelg, list):
        itc_inelg = [itc_inelg] if isinstance(itc_inelg, dict) else []
        print(f"[DEBUG] Converted itc_inelg to list: {len(itc_inelg)} items")
    summary_inelg = {}
    for item in itc_inelg:
        if not isinstance(item, dict):
            print("[DEBUG] Skipping non-dict item in itc_inelg")
            continue
        sub_type = item.get("ty", "")
        if isinstance(sub_type, str):
            sub_type = sub_type.strip()
            print(f"[DEBUG] Processing itc_inelg item with type: {sub_type}")
        row = {
            "Tax Period": tax_period,
            "Integrated Tax": get_numeric_value(item, "iamt"),
            "Central Tax": get_numeric_value(item, "camt"),
            "State/UT Tax": get_numeric_value(item, "samt"),
            "Cess": get_numeric_value(item, "csamt")
        }
        key = f"ITC-inelg-{sub_type}"
        if key in extracted_data:
            extracted_data[key].append(row)
            print(f"[DEBUG] Added row to {key}: {row}")
        if tax_period not in summary_inelg:
            summary_inelg[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_inelg[tax_period]["Integrated Tax"] += row["Integrated Tax"]
        summary_inelg[tax_period]["Central Tax"] += row["Central Tax"]
        summary_inelg[tax_period]["State/UT Tax"] += row["State/UT Tax"]
        summary_inelg[tax_period]["Cess"] += row["Cess"]

    for sp in summary_inelg.values():
        extracted_data["ITC-Ineligible"].append(sp)
        print(f"[DEBUG] Added summary row to ITC-Ineligible: {sp}")

    # Add default rows for empty sheets
    for sheet_name in extracted_data:
        if not extracted_data[sheet_name]:
            print(f"[DEBUG] No data for {sheet_name} in tax period {tax_period}, adding default row")
            default_row = {
                "Tax Period": tax_period,
                "Integrated Tax": 0.00,
                "Central Tax": 0.00,
                "State/UT Tax": 0.00,
                "Cess": 0.00
            }
            extracted_data[sheet_name].append(default_row)
            print(f"[DEBUG] Added default row to {sheet_name}: {default_row}")

    print("[DEBUG] Finished extracting section 4 data")
    return extracted_data

def extract_section_5_1(data):
    print("[DEBUG] Extracting section 5.1 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    intr_ltfee = r3b_data.get("intr_ltfee", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    extracted_data = {
        "INTR-paid": [],
        "Late-fee": []
    }
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    intr_details = intr_ltfee.get("intr_details", {})
    if isinstance(intr_details, dict):
        intr_details = [intr_details]
    elif not isinstance(intr_details, list):
        intr_details = []
        print("[DEBUG] intr_details is not a list or dict, using empty list")
    for item in intr_details:
        if not isinstance(item, dict):
            print("[DEBUG] Skipping non-dict item in intr_details")
            continue
        row = {
            "Tax Period": tax_period,
            "Integrated Tax": get_numeric_value(item, "iamt"),
            "Central Tax": get_numeric_value(item, "camt"),
            "State/UT Tax": get_numeric_value(item, "samt"),
            "Cess": get_numeric_value(item, "csamt")
        }
        extracted_data["INTR-paid"].append(row)
        print(f"[DEBUG] Added row to INTR-paid: {row}")

    ltfee_details = intr_ltfee.get("ltfee_details", {})
    if isinstance(ltfee_details, dict):
        ltfee_details = [ltfee_details]
    elif not isinstance(ltfee_details, list):
        ltfee_details = []
        print("[DEBUG] ltfee_details is not a list or dict, using empty list")
    for item in ltfee_details:
        if not isinstance(item, dict):
            print("[DEBUG] Skipping non-dict item in ltfee_details")
            continue
        row = {
            "Tax Period": tax_period,
            "Integrated Tax": get_numeric_value(item, "iamt"),
            "Central Tax": get_numeric_value(item, "camt"),
            "State/UT Tax": get_numeric_value(item, "samt"),
            "Cess": get_numeric_value(item, "csamt")
        }
        extracted_data["Late-fee"].append(row)
        print(f"[DEBUG] Added row to Late-fee: {row}")

    # Add default rows for empty sheets
    for sheet_name in extracted_data:
        if not extracted_data[sheet_name]:
            print(f"[DEBUG] No data for {sheet_name} in tax period {tax_period}, adding default row")
            default_row = {
                "Tax Period": tax_period,
                "Integrated Tax": 0.00,
                "Central Tax": 0.00,
                "State/UT Tax": 0.00,
                "Cess": 0.00
            }
            extracted_data[sheet_name].append(default_row)
            print(f"[DEBUG] Added default row to {sheet_name}: {default_row}")

    print("[DEBUG] Finished extracting section 5.1 data")
    return extracted_data

def extract_section_6(data):
    print("[DEBUG] Extracting section 6 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    tt_val = r3b_data.get("tt_val", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    extracted_data = {"Tax-Pay": []}
    row = {
        "Tax Period": tax_period,
        "Tax-by-ITC": get_numeric_value(tt_val, "tt_itc_pd"),
        "Tax-by-Cash": get_numeric_value(tt_val, "tt_csh_pd")
    }
    extracted_data["Tax-Pay"].append(row)
    print(f"[DEBUG] Extracted row for Tax-Pay: {row}")

    # Add default row if no data
    if not extracted_data["Tax-Pay"] or all(val == 0.00 for key, val in row.items() if key != "Tax Period"):
        print(f"[DEBUG] No valid data for Tax-Pay in tax period {tax_period}, ensuring default row")
        extracted_data["Tax-Pay"] = [{
            "Tax Period": tax_period,
            "Tax-by-ITC": 0.00,
            "Tax-by-Cash": 0.00
        }]
        print(f"[DEBUG] Set default row for Tax-Pay: {extracted_data['Tax-Pay'][0]}")

    print("[DEBUG] Finished extracting section 6 data")
    return extracted_data

def extract_section_6_1(data):
    print("[DEBUG] Extracting section 6.1 data...")
    tax_data = data.get("taxpayable", {}).get("data", {}).get("returnsDbCdredList", {})
    ret_period = data.get("data", {}).get("r3b", {}).get("ret_period", "")
    tax_period = get_tax_period(ret_period)

    extracted_data = {
        "6.1a1": [], "6.1a2": [], "6.1a3": [], "6.1a5": [], "6.1a6": [], "6.1a7": [],
        "6.1a41": [], "6.1a42": [], "6.1a43": [], "6.1a44": [],
        "6.1b1": [], "6.1b2": [], "6.1b3": [], "6.1b5": [], "6.1b6": [], "6.1b7": [],
        "6.1b41": [], "6.1b42": [], "6.1b43": [], "6.1b44": []
    }
    print(f"[DEBUG] Initialized extracted_data with keys: {list(extracted_data.keys())}")

    # Add default rows if no taxpayable data
    if not tax_data:
        print(f"[DEBUG] No taxpayable data for tax period {tax_period}, adding default rows")
        default_row = {
            "Tax Period": tax_period,
            "Integrated Tax": 0.00,
            "Central Tax": 0.00,
            "State/UT Tax": 0.00,
            "Cess": 0.00
        }
        for key in extracted_data:
            extracted_data[key].append(default_row)
            print(f"[DEBUG] Added default row to {key}: {default_row}")
    else:
        # Helper function to extract tax fields
        def extract_tax_fields(item, field):
            row = {
                "Tax Period": tax_period,
                "Integrated Tax": get_numeric_value(item.get("igst", {}), field),
                "Central Tax": get_numeric_value(item.get("cgst", {}), field),
                "State/UT Tax": get_numeric_value(item.get("sgst", {}), field),
                "Cess": get_numeric_value(item.get("cess", {}), field)
            }
            print(f"[DEBUG] Extracted row for field {field}: {row}")
            return row

        # Process tax_pay for 6.1a1, 6.1a6, 6.1a7 (trancd: 30002) and 6.1b1, 6.1b6, 6.1b7 (trancd: 30003)
        tax_pay = tax_data.get("tax_pay", [])
        if not isinstance(tax_pay, list):
            tax_pay = [tax_pay] if isinstance(tax_pay, dict) else []
            print(f"[DEBUG] Converted tax_pay to list: {len(tax_pay)} items")
        for item in tax_pay:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping tax_pay item: not dict")
                continue
            trancd = item.get("trancd")
            if trancd == 30002:
                # 6.1a1: Total Tax Payable
                extracted_data["6.1a1"].append(extract_tax_fields(item, "tx"))
                # 6.1a6: Interest Paid by Cash
                extracted_data["6.1a6"].append(extract_tax_fields(item, "intr"))
                # 6.1a7: Late Fee Paid by Cash
                extracted_data["6.1a7"].append(extract_tax_fields(item, "fee"))
            elif trancd == 30003:
                # 6.1b1: Total Tax Payable
                extracted_data["6.1b1"].append(extract_tax_fields(item, "tx"))
                # 6.1b6: Interest Paid by Cash
                extracted_data["6.1b6"].append(extract_tax_fields(item, "intr"))
                # 6.1b7: Late Fee Paid by Cash
                extracted_data["6.1b7"].append(extract_tax_fields(item, "fee"))
            else:
                print(f"[DEBUG] Skipping tax_pay item: trancd {trancd} not 30002 or 30003")

        # Process pd_by_nls for 6.1a2 (trancd: 30002) and 6.1b2 (trancd: 30003)
        pd_by_nls = tax_data.get("tax_paid", {}).get("pd_by_nls", [])
        if not isinstance(pd_by_nls, list):
            pd_by_nls = [pd_by_nls] if isinstance(pd_by_nls, dict) else []
            print(f"[DEBUG] Converted pd_by_nls to list: {len(pd_by_nls)} items")
        for item in pd_by_nls:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping pd_by_nls item: not dict")
                continue
            trancd = item.get("trancd")
            if trancd == 30002:
                extracted_data["6.1a2"].append(extract_tax_fields(item, "tx"))
            elif trancd == 30003:
                extracted_data["6.1b2"].append(extract_tax_fields(item, "tx"))
            else:
                print(f"[DEBUG] Skipping pd_by_nls item: trancd {trancd} not 30002 or 30003")

        # Process net_tax_pay for 6.1a3 (trancd: 30002) and 6.1b3 (trancd: 30003)
        net_tax_pay = tax_data.get("net_tax_pay", [])
        if not isinstance(net_tax_pay, list):
            net_tax_pay = [net_tax_pay] if isinstance(net_tax_pay, dict) else []
            print(f"[DEBUG] Converted net_tax_pay to list: {len(net_tax_pay)} items")
        for item in net_tax_pay:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping net_tax_pay item: not dict")
                continue
            trancd = item.get("trancd")
            if trancd == 30002:
                extracted_data["6.1a3"].append(extract_tax_fields(item, "tx"))
            elif trancd == 30003:
                extracted_data["6.1b3"].append(extract_tax_fields(item, "tx"))
            else:
                print(f"[DEBUG] Skipping net_tax_pay item: trancd {trancd} not 30002 or 30003")

        # Process pd_by_cash for 6.1a5 (trancd: 30002) and 6.1b5 (trancd: 30003)
        pd_by_cash = tax_data.get("tax_paid", {}).get("pd_by_cash", [])
        if not isinstance(pd_by_cash, list):
            pd_by_cash = [pd_by_cash] if isinstance(pd_by_cash, dict) else []
            print(f"[DEBUG] Converted pd_by_cash to list: {len(pd_by_cash)} items")
        for item in pd_by_cash:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping pd_by_cash item: not dict")
                continue
            trancd = item.get("trancd")
            if trancd == 30002:
                extracted_data["6.1a5"].append(extract_tax_fields(item, "tx"))
            elif trancd == 30003:
                extracted_data["6.1b5"].append(extract_tax_fields(item, "tx"))
            else:
                print(f"[DEBUG] Skipping pd_by_cash item: trancd {trancd} not 30002 or 30003")

        # Process pd_by_itc for 6.1a41–6.1a44 (trancd: 30002) and 6.1b41–6.1b44 (trancd: 30003)
        pd_by_itc = tax_data.get("tax_paid", {}).get("pd_by_itc", [])
        if not isinstance(pd_by_itc, list):
            pd_by_itc = [pd_by_itc] if isinstance(pd_by_itc, dict) else []
            print(f"[DEBUG] Converted pd_by_itc to list: {len(pd_by_itc)} items")
        for item in pd_by_itc:
            if not isinstance(item, dict):
                print(f"[DEBUG] Skipping pd_by_itc item: not dict")
                continue
            trancd = item.get("trancd")
            if trancd not in [30002, 30003]:
                print(f"[DEBUG] Skipping pd_by_itc item: trancd {trancd} not 30002 or 30003")
                continue
            row_igst = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
            row_cgst = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
            row_sgst = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
            row_cess = {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
            for key, val in item.items():
                if key in ["debit_id", "liab_id", "trancd", "trandate"]:
                    continue
                parts = key.split("_")
                if len(parts) != 3 or parts[2] != "amt":
                    print(f"[DEBUG] Skipping invalid ITC key: {key}")
                    continue
                prefix, middle, _ = parts
                value = parse_number(val, float_2dec=True)
                print(f"[DEBUG] Processing ITC key: {key}, value: {value}")
                if middle.lower() == "igst":
                    if prefix.lower() == "igst":
                        row_igst["Integrated Tax"] += value
                    elif prefix.lower() == "cgst":
                        row_igst["Central Tax"] += value
                    elif prefix.lower() == "sgst":
                        row_igst["State/UT Tax"] += value
                    elif prefix.lower() == "cess":
                        row_igst["Cess"] += value
                elif middle.lower() == "cgst":
                    if prefix.lower() == "igst":
                        row_cgst["Integrated Tax"] += value
                    elif prefix.lower() == "cgst":
                        row_cgst["Central Tax"] += value
                    elif prefix.lower() == "sgst":
                        row_cgst["State/UT Tax"] += value
                    elif prefix.lower() == "cess":
                        row_cgst["Cess"] += value
                elif middle.lower() == "sgst":
                    if prefix.lower() == "igst":
                        row_sgst["Integrated Tax"] += value
                    elif prefix.lower() == "cgst":
                        row_sgst["Central Tax"] += value
                    elif prefix.lower() == "sgst":
                        row_sgst["State/UT Tax"] += value
                    elif prefix.lower() == "cess":
                        row_sgst["Cess"] += value
                elif middle.lower() == "cess":
                    if prefix.lower() == "igst":
                        row_cess["Integrated Tax"] += value
                    elif prefix.lower() == "cgst":
                        row_cess["Central Tax"] += value
                    elif prefix.lower() == "sgst":
                        row_cess["State/UT Tax"] += value
                    elif prefix.lower() == "cess":
                        row_cess["Cess"] += value
            if trancd == 30002:
                extracted_data["6.1a41"].append(row_igst)
                extracted_data["6.1a42"].append(row_cgst)
                extracted_data["6.1a43"].append(row_sgst)
                extracted_data["6.1a44"].append(row_cess)
                print(f"[DEBUG] Added ITC rows (trancd 30002): igst={row_igst}, cgst={row_cgst}, sgst={row_sgst}, cess={row_cess}")
            elif trancd == 30003:
                extracted_data["6.1b41"].append(row_igst)
                extracted_data["6.1b42"].append(row_cgst)
                extracted_data["6.1b43"].append(row_sgst)
                extracted_data["6.1b44"].append(row_cess)
                print(f"[DEBUG] Added ITC rows (trancd 30003): igst={row_igst}, cgst={row_cgst}, sgst={row_sgst}, cess={row_cess}")

        # Add default rows for empty subsections
        default_row = {
            "Tax Period": tax_period,
            "Integrated Tax": 0.00,
            "Central Tax": 0.00,
            "State/UT Tax": 0.00,
            "Cess": 0.00
        }
        for key in extracted_data:
            if not extracted_data[key]:
                print(f"[DEBUG] No data for {key} in tax period {tax_period}, adding default row")
                extracted_data[key].append(default_row)
                print(f"[DEBUG] Added default row to {key}: {default_row}")

    print("[DEBUG] Finished extracting section 6.1 data")
    return extracted_data

def create_excel_report(data_dict, save_path, template_path=None):
    print("[DEBUG] Starting Excel report creation...")
    if template_path and os.path.exists(template_path):
        print(f"[DEBUG] Loading template from {template_path}")
        wb = load_workbook(template_path)
    else:
        print("[DEBUG] Creating new workbook")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            print("[DEBUG] Removed default 'Sheet'")

    section_titles = {
        "OSUP-Detail": "3.1A - Outward taxable supplies (other than zero rated, nil rated and exempted)",
        "OSUP-Zero": "3.1B - Outward taxable supplies (zero rated)",
        "OSUP-Nil,Exmp": "3.1C - Other outward supplies (nil rated, exempted)",
        "ISUP-Rev": "3.1D - Inward supplies (liable to reverse charge)",
        "OSUP-NonGST": "3.1E - Non-GST outward supplies",
        "InterSUP-Unreg": "3.2A - Inter state - supplies made to Unregistered Persons",
        "InterSUP-Comp": "3.2B - Inter state - Supplies made to Composition Taxable Persons",
        "InterSUP-UIN": "3.2C - Inter state - Supplies made to UIN holders",
        "ITC-Available": "4A - ITC Available (whether in full or part)",
        "ITC-avl-IMPG": "4A1 - ITC Available_Import of goods",
        "ITC-avl-IMPS": "4A2 - ITC Available_Import of services",
        "ITC-avl-ISRC": "4A3 - ITC Available_Inward supplies liable to reverse charge (other than 1 & 2 above)",
        "ITC-avl-ISD": "4A4 - ITC Available_Inward supplies from ISD",
        "ITC-avl-OTH": "4A5 - ITC Available_All other ITC",
        "ITC-Reversed": "4B - ITC Reversed",
        "ITC-rev-RUL": "4B1 - ITC Reversed_As per rules 38,42 & 43 of CGST Rules and section17(5)",
        "ITC-rev-OTH": "4B2 - ITC Reversed_Others",
        "Net-ITC": "4C - Net ITC Available (A-B)",
        "ITC-Ineligible": "4D - Other Details",
        "ITC-inelg-RUL": "4D1 - Other Details_ITC reclaimed which was reversed under Table 4B2 in earlier tax period",
        "ITC-inelg-OTH": "4D2 - Other Details_Ineligible ITC under section 16(4) & ITC restricted due to PoS rules",
        "INTR-paid": "5.1A - Interest Paid",
        "Late-fee": "5.1B - Late fee",
        "Tax-Pay": "6 - Payment of Tax",
        "6.1a1": "6.1A1 - Payment of Tax - Total Tax Payable (Other than reverse charge)",
        "6.1a2": "6.1A2 - Payment of Tax - Adjustment of Negative Liability (Other than reverse charge)",
        "6.1a3": "6.1A3 - Payment of Tax - Net Tax Payable (Other than reverse charge)",
        "6.1a5": "6.1A5 - Payment of Tax - Tax Paid in Cash (Other than reverse charge)",
        "6.1a6": "6.1A6 - Payment of Tax - Interest Paid by Cash (Other than reverse charge)",
        "6.1a7": "6.1A7 - Payment of Tax - Late Fee Paid by Cash (Other than reverse charge)",
        "6.1a41": "6.1A41 - Payment of Tax - Tax Paid by ITC - IGST (Other than reverse charge)",
        "6.1a42": "6.1A42 - Payment of Tax - Tax Paid by ITC - CGST (Other than reverse charge)",
        "6.1a43": "6.1A43 - Payment of Tax - Tax Paid by ITC - SGST (Other than reverse charge)",
        "6.1a44": "6.1A44 - Payment of Tax - Tax Paid by ITC - Cess (Other than reverse charge)",
        "6.1b1": "6.1B1 - Payment of Tax - Total Tax Payable (Reverse charge)",
        "6.1b2": "6.1B2 - Payment of Tax - Adjustment of Negative Liability (Reverse charge)",
        "6.1b3": "6.1B3 - Payment of Tax - Net Tax Payable (Reverse charge)",
        "6.1b5": "6.1B5 - Payment of Tax - Tax Paid in Cash (Reverse charge)",
        "6.1b6": "6.1B6 - Payment of Tax - Interest Paid by Cash (Reverse charge)",
        "6.1b7": "6.1B7 - Payment of Tax - Late Fee Paid by Cash (Reverse charge)",
        "6.1b41": "6.1B41 - Payment of Tax - Tax Paid by ITC - IGST (Reverse charge)",
        "6.1b42": "6.1B42 - Payment of Tax - Tax Paid by ITC - CGST (Reverse charge)",
        "6.1b43": "6.1B43 - Payment of Tax - Tax Paid by ITC - SGST (Reverse charge)",
        "6.1b44": "6.1B44 - Payment of Tax - Tax Paid by ITC - Cess (Reverse charge)",
    }

    sheet_names = {
        "OSUP-Detail": "3B-OSUP-Detail",
        "OSUP-Zero": "3B-OSUP-Zero",
        "OSUP-Nil,Exmp": "3B-OSUP-Nil,Exmp",
        "ISUP-Rev": "3B-ISUP-Rev",
        "OSUP-NonGST": "3B-OSUP-NonGST",
        "InterSUP-Unreg": "3B-InterSUP-Unreg",
        "InterSUP-Comp": "3B-InterSUP-Comp",
        "InterSUP-UIN": "3B-InterSUP-UIN",
        "ITC-Available": "3B-ITC-Available",
        "ITC-avl-IMPG": "3B-ITC-avl-IMPG",
        "ITC-avl-IMPS": "3B-ITC-avl-IMPS",
        "ITC-avl-ISRC": "3B-ITC-avl-ISRC",
        "ITC-avl-ISD": "3B-ITC-avl-ISD",
        "ITC-avl-OTH": "3B-ITC-avl-OTH",
        "ITC-Reversed": "3B-ITC-Reversed",
        "ITC-rev-RUL": "3B-ITC-rev-RUL",
        "ITC-rev-OTH": "3B-ITC-rev-OTH",
        "Net-ITC": "3B-Net-ITC",
        "ITC-Ineligible": "3B-ITC-Ineligible",
        "ITC-inelg-RUL": "3B-ITC-inelg-RUL",
        "ITC-inelg-OTH": "3B-ITC-inelg-OTH",
        "INTR-paid": "3B-INTR-paid",
        "Late-fee": "3B-Late-fee",
        "Tax-Pay": "3B-Tax-Pay",
        "6.1a1": "3B-TaxPay_TotTax-OthRC",
        "6.1a2": "3B-TaxPay_AdjNL-OthRC",
        "6.1a3": "3B-TaxPay_NetTax-OthRC",
        "6.1a5": "3B-TaxPay_pdby_Cash-OthRC",
        "6.1a6": "3B-TaxPay_Int_pdby_Cash-OthRC",
        "6.1a7": "3B-TaxPay_LateFee_pdby_Cash-OthRC",
        "6.1a41": "3B-TaxPay_ITC_IGST-OthRC",
        "6.1a42": "3B-TaxPay_ITC_CGST-OthRC",
        "6.1a43": "3B-TaxPay_ITC_SGST-OthRC",
        "6.1a44": "3B-TaxPay_ITC_Cess-OthRC",
        "6.1b1": "3B-TaxPay_TotTax-RC",
        "6.1b2": "3B-TaxPay_AdjNL-RC",
        "6.1b3": "3B-TaxPay_NetTax-RC",
        "6.1b5": "3B-TaxPay_pdby_Cash-RC",
        "6.1b6": "3B-TaxPay_Int_pdby_Cash-RC",
        "6.1b7": "3B-TaxPay_LateFee_pdby_Cash-RC",
        "6.1b41": "3B-TaxPay_ITC_IGST-RC",
        "6.1b42": "3B-TaxPay_ITC_CGST-RC",
        "6.1b43": "3B-TaxPay_ITC_SGST-RC",
        "6.1b44": "3B-TaxPay_ITC_Cess-RC",
    }

    column_headers = {
        "OSUP-Detail": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-Zero": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-Nil,Exmp": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ISUP-Rev": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-NonGST": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "InterSUP-Unreg": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "InterSUP-Comp": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "InterSUP-UIN": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "ITC-Available": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-IMPG": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-IMPS": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-ISRC": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-ISD": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-Reversed": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-rev-RUL": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-rev-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Net-ITC": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-Ineligible": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-inelg-RUL": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-inelg-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "INTR-paid": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Late-fee": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Tax-Pay": ["Tax Period", "Tax-by-ITC", "Tax-by-Cash"],
        "6.1a1": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a2": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a3": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a5": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a6": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a7": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a41": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a42": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a43": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a44": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b1": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b2": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b3": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b5": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b6": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b7": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b41": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b42": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b43": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b44": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    }

    indian_format = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    column_formats = {
        "OSUP-Detail": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "OSUP-Zero": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "OSUP-Nil,Exmp": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ISUP-Rev": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "OSUP-NonGST": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "InterSUP-Unreg": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format},
        "InterSUP-Comp": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format},
        "InterSUP-UIN": {"Tax Period": "General", "Total taxable value": indian_format, "Integrated Tax": indian_format},
        "ITC-Available": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-avl-IMPG": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-avl-IMPS": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-avl-ISRC": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-avl-ISD": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-avl-OTH": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-Reversed": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-rev-RUL": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-rev-OTH": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "Net-ITC": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-Ineligible": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-inelg-RUL": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "ITC-inelg-OTH": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "INTR-paid": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "Late-fee": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "Tax-Pay": {"Tax Period": "General", "Tax-by-ITC": indian_format, "Tax-by-Cash": indian_format},
        "6.1a1": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a2": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a3": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a5": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a6": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a7": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a41": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a42": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a43": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1a44": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b1": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b2": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b3": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b5": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b6": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b7": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b41": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b42": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b43": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
        "6.1b44": {"Tax Period": "General", "Integrated Tax": indian_format, "Central Tax": indian_format, "State/UT Tax": indian_format, "Cess": indian_format},
    }

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    def sheet_has_valid_data(rows, numeric_headers):
        print(f"[DEBUG] Checking if sheet has valid data, numeric_headers: {numeric_headers}")
        for row in rows:
            for header in numeric_headers:
                try:
                    value = float(row.get(header, 0))
                    if value != 0:
                        print(f"[DEBUG] Found non-zero value for {header}: {value}")
                        return True
                except (ValueError, TypeError):
                    print(f"[DEBUG] Invalid value for {header}: {row.get(header)}")
                    continue
        print("[DEBUG] No non-zero numeric values found, skipping sheet")
        return False

    print("[DEBUG] Generating sheets...")
    for sheet_key, rows in data_dict.items():
        print(f"[DEBUG] Processing sheet: {sheet_key}, rows: {len(rows)}")
        numeric_headers = [h for h in column_formats.get(sheet_key, {}).keys() if h != "Tax Period"]
        if not sheet_has_valid_data(rows, numeric_headers):
            print(f"[DEBUG] Skipping sheet {sheet_key} due to no non-zero numeric data")
            continue

        new_sheet_name = sheet_names.get(sheet_key, "3B-" + sheet_key)
        print(f"[DEBUG] Creating sheet: {new_sheet_name}")
        if new_sheet_name in wb.sheetnames:
            wb.remove(wb[new_sheet_name])
            print(f"[DEBUG] Removed existing sheet: {new_sheet_name}")

        ws = wb.create_sheet(new_sheet_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(column_headers[sheet_key]))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = section_titles[sheet_key]
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        print(f"[DEBUG] Set sheet title: {section_titles[sheet_key]}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        print(f"[DEBUG] Added headers: {column_headers[sheet_key]}")

        for row_idx, row_data in enumerate(rows, start=3):
            for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
                cell_value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if col_name in column_formats[sheet_key]:
                    cell.number_format = column_formats[sheet_key][col_name]
        print(f"[DEBUG] Populated data rows for {new_sheet_name}")

        for col_idx, col_name in enumerate(column_headers[sheet_key], start=1):
            col_letter = get_column_letter(col_idx)
            max_length = len(col_name)
            for row in rows:
                max_length = max(max_length, len(str(row.get(col_name, ""))))
            ws.column_dimensions[col_letter].width = max(15, max_length + 2)
        print(f"[DEBUG] Adjusted column widths for {new_sheet_name}")

        print(f"[DEBUG] Created sheet: {new_sheet_name}")

    print(f"[DEBUG] Saving workbook to {save_path}...")
    wb.save(save_path)
    print(f"[DEBUG] Workbook saved successfully to {save_path}")
    print("[DEBUG] Finished Excel report creation")
    return f"✅ Successfully saved Excel file: {save_path}"

def process_gstr3b(json_files, template_path, save_path):
    print("[DEBUG] Starting GSTR-3B processing...")
    combined_data = {
        "OSUP-Detail": [],
        "OSUP-Zero": [],
        "OSUP-Nil,Exmp": [],
        "ISUP-Rev": [],
        "OSUP-NonGST": [],
        "InterSUP-Unreg": [],
        "InterSUP-Comp": [],
        "InterSUP-UIN": [],
        "ITC-Available": [],
        "ITC-avl-IMPG": [],
        "ITC-avl-IMPS": [],
        "ITC-avl-ISRC": [],
        "ITC-avl-ISD": [],
        "ITC-avl-OTH": [],
        "ITC-Reversed": [],
        "ITC-rev-RUL": [],
        "ITC-rev-OTH": [],
        "Net-ITC": [],
        "ITC-Ineligible": [],
        "ITC-inelg-RUL": [],
        "ITC-inelg-OTH": [],
        "INTR-paid": [],
        "Late-fee": [],
        "Tax-Pay": [],
        "6.1a1": [],
        "6.1a2": [],
        "6.1a3": [],
        "6.1a5": [],
        "6.1a6": [],
        "6.1a7": [],
        "6.1a41": [],
        "6.1a42": [],
        "6.1a43": [],
        "6.1a44": [],
        "6.1b1": [],
        "6.1b2": [],
        "6.1b3": [],
        "6.1b5": [],
        "6.1b6": [],
        "6.1b7": [],
        "6.1b41": [],
        "6.1b42": [],
        "6.1b43": [],
        "6.1b44": []
    }
    print(f"[DEBUG] Initialized combined_data with keys: {list(combined_data.keys())}")

    print("[DEBUG] Processing JSON files...")
    for file_path in json_files:
        print(f"[DEBUG] Loading JSON file: {file_path}")
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            print(f"[DEBUG] Loaded JSON data from {file_path}")

            extracted_3_1 = extract_section_3_1(data)
            for key, rows in extracted_3_1.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

            extracted_3_2 = extract_section_3_2(data)
            for key, rows in extracted_3_2.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

            extracted_4 = extract_section_4(data)
            for key, rows in extracted_4.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

            extracted_5_1 = extract_section_5_1(data)
            for key, rows in extracted_5_1.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

            extracted_6 = extract_section_6(data)
            for key, rows in extracted_6.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

            extracted_6_1 = extract_section_6_1(data)
            for key, rows in extracted_6_1.items():
                combined_data[key].extend(rows)
                print(f"[DEBUG] Extended {key} with {len(rows)} rows")

    print("[DEBUG] Finished processing JSON files")

    print("[DEBUG] Sorting data...")
    financial_order = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
    for key in combined_data:
        print(f"[DEBUG] Sorting section: {key}")
        combined_data[key].sort(key=lambda x: financial_order.index(x["Tax Period"]) if x["Tax Period"] in financial_order else 999)
        print(f"[DEBUG] Sorted {key} with {len(combined_data[key])} rows")
    print("[DEBUG] Data sorting completed")

    print("[DEBUG] Generating Excel report...")
    result = create_excel_report(combined_data, save_path, template_path)
    print("[DEBUG] GSTR-3B processing completed")
    return result