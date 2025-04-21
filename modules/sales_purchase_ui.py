# sales_purchase_ui.py

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import Workbook, load_workbook

from telemetry import send_event
from sales_processor import process_excel_data
from purchase_processor import process_purchase_data

class SalesPurchaseProcessorUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Process Sales / Purchase")
        self.root.geometry("500x480")
        self.sales_files = []
        self.purchase_files = []
        self.template_file = None
        self.base_height = 450

        tk.Label(root, text="Process Sales / Purchase", font=("Arial", 16, "bold")).pack(pady=5)
        main_frame = tk.Frame(root)
        main_frame.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)

        # Sales Section
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(left_frame, text="Sales Ledger Excel Files", font=("Arial", 10, "bold")).pack()
        self.sales_tree = ttk.Treeview(left_frame, columns=("File Name", "Branch Code"),
                                       show="headings", height=13)
        self.sales_tree.heading("File Name", text="File Name")
        self.sales_tree.heading("Branch Code", text="Branch Code")
        self.sales_tree.column("File Name", width=140, stretch=True)
        self.sales_tree.column("Branch Code", width=85, stretch=True)
        self.sales_tree.pack(pady=2, fill=tk.Y, expand=True)
        self.sales_tree.bind("<Double-1>", self.edit_sales_branch_code)
        btn_frame = tk.Frame(left_frame)
        btn_frame.pack(pady=5)
        tk.Button(btn_frame, text="+ Add", command=self.add_sales_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="- Remove", command=self.delete_sales_file).pack(side=tk.LEFT, padx=5)

        # Purchase Section
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tk.Label(right_frame, text="Purchase Register Excel Files", font=("Arial", 10, "bold")).pack()
        self.purchase_tree = ttk.Treeview(right_frame, columns=("File Name", "Branch Code"),
                                          show="headings", height=13)
        self.purchase_tree.heading("File Name", text="File Name")
        self.purchase_tree.heading("Branch Code", text="Branch Code")
        self.purchase_tree.column("File Name", width=140, stretch=False)
        self.purchase_tree.column("Branch Code", width=85, stretch=False)
        self.purchase_tree.pack(pady=2, fill=tk.Y, expand=True)
        self.purchase_tree.bind("<Double-1>", self.edit_purchase_branch_code)
        btn_frame = tk.Frame(right_frame)
        btn_frame.pack(pady=5)
        tk.Button(btn_frame, text="+ Add", command=self.add_purchase_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="- Remove", command=self.delete_purchase_file).pack(side=tk.LEFT, padx=5)

        # Template File Section
        template_frame = tk.Frame(root)
        template_frame.pack(pady=5)
        tk.Label(template_frame, text="Template Excel File (Optional):").pack(side=tk.LEFT)
        self.template_label = tk.Label(template_frame, text="No file selected")
        self.template_label.pack(side=tk.LEFT, padx=5)
        tk.Button(template_frame, text="Select", command=self.select_template).pack(side=tk.LEFT, padx=2)
        tk.Button(template_frame, text="Clear", command=self.clear_template).pack(side=tk.LEFT, padx=2)

        # Warning Frame
        self.warning_frame = tk.Frame(root, borderwidth=1, relief="solid")
        self.warning_title = tk.Label(self.warning_frame, text="Warning!", fg="red",
                                      font=("Arial", 10, "underline"))
        self.warning_text = tk.Label(self.warning_frame, text="", fg="red",
                                     justify=tk.LEFT, wraplength=450)
        self.ignore_var = tk.BooleanVar()
        self.ignore_check = tk.Checkbutton(self.warning_frame, text="Ignore Warning",
                                           variable=self.ignore_var,
                                           command=self.update_process_button)

        # Process Button
        self.process_btn = tk.Button(root, text="Process Sales / Purchase", font=("Arial", 12),
                                     command=self.process_files, state=tk.DISABLED, bg="light grey")
        self.process_btn.pack(pady=10)

        self.update_process_button()

    def add_sales_file(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx *.xls")])
        for file in files:
            if file not in [f[0] for f in self.sales_files]:
                self.sales_files.append((file, "Default"))
                self.sales_tree.insert("", tk.END,
                                       values=(os.path.basename(file), "Default"))
        self.update_process_button()

    def delete_sales_file(self):
        selected = self.sales_tree.selection()
        for item in selected:
            idx = self.sales_tree.index(item)
            self.sales_files.pop(idx)
            self.sales_tree.delete(item)
        self.update_process_button()

    def add_purchase_file(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx *.xls")])
        for file in files:
            if file not in [f[0] for f in self.purchase_files]:
                self.purchase_files.append((file, "Default"))
                self.purchase_tree.insert("", tk.END,
                                          values=(os.path.basename(file), "Default"))
        self.update_process_button()

    def delete_purchase_file(self):
        selected = self.purchase_tree.selection()
        for item in selected:
            idx = self.purchase_tree.index(item)
            self.purchase_files.pop(idx)
            self.purchase_tree.delete(item)
        self.update_process_button()

    def edit_sales_branch_code(self, event):
        item = self.sales_tree.identify_row(event.y)
        column = self.sales_tree.identify_column(event.x)
        if not item or column != "#2":
            return
        idx = self.sales_tree.index(item)
        old = self.sales_files[idx][1]
        entry = tk.Entry(self.sales_tree)
        entry.insert(0, old)
        x, y, width, _ = self.sales_tree.bbox(item, column)
        entry.place(x=x, y=y, width=self.sales_tree.column("Branch Code")["width"])
        entry.focus_set()
        def save(evt):
            new = entry.get().strip() or "Default"
            self.sales_files[idx] = (self.sales_files[idx][0], new)
            self.sales_tree.item(item, values=(os.path.basename(self.sales_files[idx][0]), new))
            entry.destroy()
            self.update_process_button()
        entry.bind("<Return>", save)
        entry.bind("<FocusOut>", save)

    def edit_purchase_branch_code(self, event):
        item = self.purchase_tree.identify_row(event.y)
        column = self.purchase_tree.identify_column(event.x)
        if not item or column != "#2":
            return
        idx = self.purchase_tree.index(item)
        old = self.purchase_files[idx][1]
        entry = tk.Entry(self.purchase_tree)
        entry.insert(0, old)
        x, y, width, _ = self.purchase_tree.bbox(item, column)
        entry.place(x=x, y=y, width=self.purchase_tree.column("Branch Code")["width"])
        entry.focus_set()
        def save(evt):
            new = entry.get().strip() or "Default"
            self.purchase_files[idx] = (self.purchase_files[idx][0], new)
            self.purchase_tree.item(item, values=(os.path.basename(self.purchase_files[idx][0]), new))
            entry.destroy()
            self.update_process_button()
        entry.bind("<Return>", save)
        entry.bind("<FocusOut>", save)

    def select_template(self):
        file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file:
            self.template_file = file
            self.template_label.config(text=os.path.basename(file))

    def clear_template(self):
        self.template_file = None
        self.template_label.config(text="No file selected")

    def update_process_button(self):
        has_files = bool(self.sales_files or self.purchase_files)
        missing = any(not code.strip() for _, code in self.sales_files + self.purchase_files)
        if missing:
            if not self.warning_frame.winfo_ismapped():
                self.warning_frame.pack(pady=5, padx=10, fill=tk.X)
            self.warning_title.pack(pady=2)
            self.warning_text.config(text="Warning: Branch code missing for some files.")
            self.warning_text.pack(pady=2)
            self.ignore_check.pack(pady=2)
            self.root.geometry(f"500x{self.base_height + 100}")
            if has_files and self.ignore_var.get():
                self.process_btn.config(state=tk.NORMAL, bg="light green")
            else:
                self.process_btn.config(state=tk.DISABLED, bg="light grey")
        else:
            if self.warning_frame.winfo_ismapped():
                self.warning_frame.pack_forget()
                self.root.geometry(f"500x{self.base_height}")
            if has_files:
                self.process_btn.config(state=tk.NORMAL, bg="light green")
            else:
                self.process_btn.config(state=tk.DISABLED, bg="light grey")

    def process_files(self):
        if not (self.sales_files or self.purchase_files):
            messagebox.showerror("Error", "No files selected for processing.")
            return

        missing = any(not code.strip() for _, code in self.sales_files + self.purchase_files)
        if missing and not self.ignore_var.get():
            self.update_process_button()
            return

        sales = [(f, c.strip() or "Default") for f, c in self.sales_files]
        purchase = [(f, c.strip() or "Default") for f, c in self.purchase_files]

        save_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Files", "*.xlsx")],
                                                 title="Save Sales/Purchase Report As")
        if not save_file:
            return

        self.process_btn.config(text="Processing...", state=tk.DISABLED, bg="light grey")
        self.root.update()

        try:
            if self.template_file:
                wb = load_workbook(self.template_file)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            if sales:
                wb = process_excel_data(sales, self.template_file, wb)
            if purchase:
                wb = process_purchase_data(purchase, self.template_file, wb)

            wb.save(save_file)

            send_event("sales_purchase_complete", {
                "sales_files": [f for f, _ in sales],
                "purchase_files": [f for f, _ in purchase],
                "output_file": save_file
            })

            messagebox.showinfo("Success", f"Report saved successfully at:\n{save_file}")
            self.sales_files.clear()
            self.purchase_files.clear()
            self.sales_tree.delete(*self.sales_tree.get_children())
            self.purchase_tree.delete(*self.purchase_tree.get_children())
            self.ignore_var.set(False)
            self.process_btn.config(text="Process Sales / Purchase")
            self.update_process_button()

        except Exception as e:
            send_event("error", {
                "module": "sales_purchase_ui",
                "error": str(e),
                "sales_files": [f for f, _ in self.sales_files],
                "purchase_files": [f for f, _ in self.purchase_files]
            })

            self.show_error_with_copy("Error", f"An error occurred:\n{str(e)}")
            self.process_btn.config(text="Process Sales / Purchase")
            self.update_process_button()

    def show_error_with_copy(self, title, message):
        print(f"ERROR: {message}")
        error_window = tk.Toplevel(self.root)
        error_window.title(title)
        error_window.geometry("400x150")
        error_window.resizable(False, False)
        tk.Label(error_window, text=message, wraplength=380, justify="left").pack(pady=10)

        def copy_error():
            import pyperclip
            pyperclip.copy(message)
            copy_button.config(text="Copied!")
            error_window.after(2000, lambda: copy_button.config(text="Copy Error"))

        copy_button = tk.Button(error_window, text="Copy Error", command=copy_error)
        copy_button.pack(side="left", padx=10, pady=10)
        tk.Button(error_window, text="OK", command=error_window.destroy).pack(side="right", padx=10, pady=10)
        error_window.transient(self.root)
        error_window.grab_set()
        self.root.wait_window(error_window)

if __name__ == "__main__":
    root = tk.Tk()
    app = SalesPurchaseProcessorUI(root)
    root.mainloop()
