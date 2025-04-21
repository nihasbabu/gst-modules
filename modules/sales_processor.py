import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import pyperclip
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Sheet titles with desired order
SECTION_TITLES = [
    ("SALE-Total", "Sales Register - Total"),
    ("SALE-Total_sws", "Sales Register - Total - Receiver wise"),
    ("SALE-B2B", "Sales Register - B2B only"),
    ("SALE-B2C", "Sales Register - B2C only"),
    ("SALE-Summary-Total", "Sales Register - Total - Summary"),
    ("SALE-Summary-B2B", "Sales Register - B2B only - Summary"),
    ("SALE-Summary-B2C", "Sales Register - B2C only - Summary"),
]

def find_header_row(worksheet):
    logging.debug("Searching for header row starting with 'Date'")
    for row in worksheet.iter_rows():
        if row[0].value == "Date":
            logging.debug(f"Found header row at row {row[0].row}")
            return row[0].row
    logging.debug("Header row not found")
    return None

def get_financial_year(date):
    logging.debug(f"Calculating financial year for date: {date}")
    if date is None:
        logging.debug("Date is None, returning None")
        return None
    if date.month >= 4:
        result = date.year
        logging.debug(f"Month >= April, financial year: {result}")
        return result
    result = date.year - 1
    logging.debug(f"Month < April, financial year: {result}")
    return result

def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    logging.debug(f"Creating or replacing sheet: {sheet_name}")
    if sheet_name in wb.sheetnames:
        logging.debug(f"Sheet {sheet_name} exists, removing it")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    logging.debug(f"Created new sheet: {sheet_name}")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.debug(f"Set sheet title: {title_text}")
    for idx, col in enumerate(columns, start=1):
        header_cell = ws.cell(row=2, column=idx, value=col)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    logging.debug(f"Added headers: {columns}")
    ws.freeze_panes = "B3"
    logging.debug("Set freeze panes at B3")
    return ws

def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    logging.debug(f"Applying formats and autofitting columns for sheet: {ws.title}")
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        logging.debug(f"Processing column: {col_name} (col {col_letter})")
        if col_format_map and col_name in col_format_map:
            logging.debug(f"Applying number format to {col_name}: {col_format_map[col_name]}")
            for row in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.number_format = col_format_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 1)
        logging.debug(f"Set column {col_letter} width to {ws.column_dimensions[col_letter].width}")
    logging.debug("Finished applying formats and autofitting")

def add_total_row(ws, columns, start_row, end_row):
    logging.debug(f"Adding total row for sheet: {ws.title}, from row {start_row} to {end_row}")
    total_row = ['Total'] + [''] * (len(columns) - 1)
    numeric_cols = ['Invoice value', 'Taxable Value', 'Integrated Tax', 'Central Tax',
                    'State/UT Tax', 'Cess', 'Addl. Cost', 'Round Off']
    logging.debug(f"Numeric columns for totals: {numeric_cols}")

    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in numeric_cols:
            logging.debug(f"Calculating total for column: {col_name}")
            total = 0
            for row in range(start_row, end_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                total += float(value) if value and isinstance(value, (int, float)) else 0
            total_row[col_idx - 1] = total
            logging.debug(f"Total for {col_name}: {total}")

    row_num = end_row + 1
    logging.debug(f"Writing total row at row {row_num}")
    for col_idx, value in enumerate(total_row, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(color="FF0000")
        if col_idx > 1 and value != '':
            cell.number_format = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    logging.debug("Total row added successfully")

def process_excel_data(input_files, template_file=None, existing_wb=None):
    logging.debug(f"Starting processing with {len(input_files)} input files")
    all_data = []
    data_by_type = {"B2B": [], "B2C": []}
    logging.debug(f"Initialized data structures: all_data, data_by_type={list(data_by_type.keys())}")

    column_mapping = {
        'GSTIN/UIN': 'GSTIN/UIN of Recipient',
        'Particulars': 'Receiver Name',
        'Voucher No': 'Invoice number',
        'Voucher No.': 'Invoice number',
        'Date': 'Invoice date',
        'Gross Total': 'Invoice value',
        'Voucher Type': 'Invoice Type',
        'Value': 'Taxable Value',
        'IGST': 'Integrated Tax',
        'CGST': 'Central Tax',
        'SGST': 'State/UT Tax',
        'Cess': 'Cess',
        'ROUND OFF': 'Round Off'
    }
    logging.debug(f"Column mapping defined: {column_mapping}")

    desired_headers = [
        'GSTIN/UIN of Recipient',
        'Receiver Name',
        'Branch',
        'Invoice number',
        'Invoice date',
        'Invoice Type',
        'Invoice value',
        'Taxable Value',
        'Integrated Tax',
        'Central Tax',
        'State/UT Tax',
        'Cess'
    ]
    logging.debug(f"Desired headers: {desired_headers}")

    additional_headers = set()
    logging.debug("Initialized additional_headers set")

    for filepath, branch_key in input_files:
        logging.debug(f"Processing file: {filepath} with branch_key: {branch_key}")
        wb = load_workbook(filepath)
        logging.debug(f"Loaded workbook: {filepath}")

        if len(wb.sheetnames) > 1:
            logging.debug(f"Multiple sheets found: {wb.sheetnames}")
            if "Sales Register" in wb.sheetnames:
                ws = wb["Sales Register"]
                logging.debug("Selected 'Sales Register' sheet")
            else:
                logging.error(f"'Sales Register' sheet not found in {filepath}")
                raise ValueError(f"Multiple sheets found but 'Sales Register' not present in {filepath}")
        else:
            ws = wb.active
            logging.debug("Single sheet found, using active sheet")
            if not ws:
                logging.error(f"No active sheet in {filepath}")
                raise ValueError(f"No active sheet found in {filepath}")

        header_row = find_header_row(ws)
        if not header_row:
            logging.error(f"Header row not found in {filepath}")
            raise ValueError(f"Could not find header row starting with 'Date' in {filepath}")
        logging.debug(f"Header row found at row {header_row}")

        headers = [cell.value for cell in ws[header_row] if cell.value is not None]
        logging.debug(f"Headers extracted: {headers}")
        additional_headers.update(h for h in headers if h not in column_mapping)
        logging.debug(f"Updated additional_headers: {additional_headers}")

        for row in ws.iter_rows(min_row=header_row + 1):
            row_data_orig = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
            logging.debug(f"Processing row: {row_data_orig}")

            if not row_data_orig.get('Date') or 'Grand Total' in str(row_data_orig.get('Particulars', '')):
                logging.debug("Skipping row: no Date or contains 'Grand Total'")
                continue

            if isinstance(row_data_orig['Date'], str):
                logging.debug(f"Date is string: {row_data_orig['Date']}")
                try:
                    row_data_orig['Date'] = datetime.datetime.strptime(row_data_orig['Date'], '%Y-%m-%d %H:%M:%S')
                    logging.debug(f"Parsed date: {row_data_orig['Date']}")
                except ValueError:
                    logging.debug("Failed to parse date, skipping row")
                    continue
            elif not isinstance(row_data_orig['Date'], datetime.datetime):
                logging.debug(f"Date is not datetime: {type(row_data_orig['Date'])}, skipping row")
                continue

            if 'ROUND OFF' in row_data_orig and 'Round Off' in row_data_orig:
                row_data_orig['Round Off'] = row_data_orig.get('ROUND OFF', 0) or row_data_orig.get('Round Off', 0)
                del row_data_orig['ROUND OFF']
                logging.debug("Consolidated 'ROUND OFF' into 'Round Off'")

            row_data = {column_mapping.get(k, k): v for k, v in row_data_orig.items()}
            logging.debug(f"Mapped row data: {row_data}")
            inv_type = row_data.get('Invoice Type')
            invoice_num = row_data.get('Invoice number')

            if not inv_type or invoice_num is None:
                logging.debug(f"Skipping row: inv_type={inv_type}, invoice_num={invoice_num}")
                continue

            row_data['Branch'] = branch_key
            logging.debug(f"Added branch_key: {branch_key}")
            all_data.append(row_data)
            logging.debug(f"Added row to all_data, total count: {len(all_data)}")
            if inv_type in data_by_type:
                data_by_type[inv_type].append(row_data)
                logging.debug(f"Added row to data_by_type[{inv_type}], count: {len(data_by_type[inv_type])}")

    logging.info(f"Total records processed: {len(all_data)}")
    logging.debug(f"Data by type: B2B={len(data_by_type['B2B'])}, B2C={len(data_by_type['B2C'])}")

    final_headers = desired_headers + [h for h in additional_headers if h not in column_mapping.keys()]
    logging.debug(f"Final headers: {final_headers}")

    for data in [all_data, data_by_type.get("B2B", []), data_by_type.get("B2C", [])]:
        logging.debug(f"Sorting data list with {len(data)} records")
        data.sort(key=lambda x: x['Invoice date'] if x['Invoice date'] is not None else datetime.datetime.min)
        logging.debug("Sorting completed")

    def sort_key(row):
        receiver = row.get('Receiver Name', '')
        date = row.get('Invoice date', datetime.datetime.min)
        logging.debug(f"Sorting row with receiver: {receiver}, date: {date}")
        if receiver.lower() in ['cash', '(cancelled )'] or not receiver:
            logging.debug(f"Receiver marked as secondary: {receiver}")
            return (1, receiver or '', date)
        logging.debug(f"Receiver marked as primary: {receiver}")
        return (0, receiver, date)

    logging.debug("Sorting all_data for SALE-Total_sws")
    all_data_sws = sorted(all_data, key=sort_key)
    logging.debug(f"Sorted all_data_sws with {len(all_data_sws)} records")

    # Use existing workbook or load template or create new
    logging.debug("Determining workbook to use")
    if existing_wb is not None:
        output_wb = existing_wb
        logging.debug("Using provided existing workbook")
    elif template_file:
        logging.debug(f"Loading template: {template_file}")
        output_wb = load_workbook(template_file)
    else:
        logging.debug("Creating new workbook")
        output_wb = Workbook()
        if 'Sheet' in output_wb.sheetnames:
            logging.debug("Removing default 'Sheet'")
            del output_wb['Sheet']

    col_format_map = {
        'Invoice date': 'DD-MM-YYYY',
        'Invoice value': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Taxable Value': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Integrated Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Central Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'State/UT Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Cess': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    }
    logging.debug(f"Column format map: {col_format_map}")

    sheets_to_create = [
        ("SALE-Total", all_data),
        ("SALE-Total_sws", all_data_sws),
        ("SALE-B2B", data_by_type.get("B2B", [])),
        ("SALE-B2C", data_by_type.get("B2C", [])),
    ]
    logging.debug(f"Sheets to create: {[s[0] for s in sheets_to_create]}")

    for sheet_name, data in sheets_to_create:
        logging.debug(f"Processing sheet: {sheet_name} with {len(data)} records")
        title = next(t for n, t in SECTION_TITLES if n == sheet_name)
        logging.debug(f"Sheet title: {title}")
        ws = create_or_replace_sheet(output_wb, sheet_name, title, final_headers)
        start_row = 3
        logging.debug(f"Starting data population at row {start_row}")
        for row_idx, row_data in enumerate(data, start=start_row):
            row = [row_data.get(header, '') if header != 'Invoice date' else
                   row_data[header].strftime('%d-%m-%Y') if row_data.get(header) else ''
                   for header in final_headers]
            ws.append(row)
            logging.debug(f"Appended row {row_idx} to {sheet_name}")
        if data:
            logging.debug(f"Adding total row for {sheet_name}")
            add_total_row(ws, final_headers, start_row, start_row + len(data) - 1)
        apply_format_and_autofit(ws, final_headers, col_format_map=col_format_map)
        logging.info(f"Created sheet {sheet_name} with {len(data)} records")

    summary_sheets = [
        ("SALE-Summary-Total", all_data),
        ("SALE-Summary-B2B", data_by_type.get("B2B", [])),
        ("SALE-Summary-B2C", data_by_type.get("B2C", []))
    ]
    logging.debug(f"Summary sheets to create: {[s[0] for s in summary_sheets]}")

    summary_headers = ['Month', 'No. of Records', 'Taxable Value', 'Integrated Tax',
                       'Central Tax', 'State/UT Tax', 'Cess']
    logging.debug(f"Summary headers: {summary_headers}")

    summary_col_format_map = {
        'Taxable Value': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Integrated Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Central Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'State/UT Tax': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;",
        'Cess': r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
    }
    logging.debug(f"Summary column format map: {summary_col_format_map}")

    months = ['April', 'May', 'June', 'July', 'August', 'September',
              'October', 'November', 'December', 'January', 'February', 'March']
    logging.debug(f"Month order for summaries: {months}")

    for sheet_name, data in summary_sheets:
        logging.debug(f"Processing summary sheet: {sheet_name} with {len(data)} records")
        title = next(t for n, t in SECTION_TITLES if n == sheet_name)
        logging.debug(f"Summary sheet title: {title}")
        ws = create_or_replace_sheet(output_wb, sheet_name, title, summary_headers)
        summary_data = {}
        invoice_numbers = set()
        logging.debug("Initialized summary_data and invoice_numbers")

        for row in data:
            date = row['Invoice date']
            invoice_num = row.get('Invoice number')
            logging.debug(f"Processing summary row: invoice_num={invoice_num}, date={date}")
            if invoice_num is None:
                logging.debug("Skipping row: no invoice number")
                continue
            if isinstance(date, datetime.datetime):
                fy = get_financial_year(date)
                month_idx = (date.month - 4) % 12
                month = months[month_idx]
                logging.debug(f"Month calculated: {month} (FY {fy})")
                if month not in summary_data:
                    summary_data[month] = {
                        'count': 0, 'Taxable Value': 0, 'Integrated Tax': 0,
                        'Central Tax': 0, 'State/UT Tax': 0, 'Cess': 0
                    }
                    logging.debug(f"Initialized summary_data for {month}")
                if invoice_num not in invoice_numbers:
                    summary_data[month]['count'] += 1
                    invoice_numbers.add(invoice_num)
                    logging.debug(f"Incremented count for {month}, unique invoice: {invoice_num}")
                for field in ['Taxable Value', 'Integrated Tax', 'Central Tax', 'State/UT Tax', 'Cess']:
                    value = row.get(field, 0) or 0
                    summary_data[month][field] += value
                    logging.debug(f"Added {field}={value} to {month}")

        start_row = 3
        row_count = 0
        logging.debug("Populating summary sheet")
        for month in months:
            if month in summary_data:
                ws.append([
                    month,
                    summary_data[month]['count'],
                    summary_data[month]['Taxable Value'],
                    summary_data[month]['Integrated Tax'],
                    summary_data[month]['Central Tax'],
                    summary_data[month]['State/UT Tax'],
                    summary_data[month]['Cess']
                ])
                row_count += 1
                logging.debug(f"Appended summary row for {month}")

        if summary_data:
            logging.debug(f"Adding total row for {sheet_name}")
            total_row = ['Total']
            for col_idx in range(1, len(summary_headers)):
                total = 0
                for row in range(start_row, start_row + row_count):
                    value = ws.cell(row=row, column=col_idx + 1).value
                    total += float(value) if value and isinstance(value, (int, float)) else 0
                total_row.append(total)
                logging.debug(f"Total for column {summary_headers[col_idx]}: {total}")
            ws.append(total_row)
            logging.debug(f"Writing total row at row {start_row + row_count}")
            for col_idx in range(1, len(summary_headers) + 1):
                cell = ws.cell(row=start_row + row_count, column=col_idx)
                cell.font = Font(color="FF0000")
                if col_idx > 1:
                    cell.number_format = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"
            logging.debug("Total row formatted")

        apply_format_and_autofit(ws, summary_headers, col_format_map=summary_col_format_map)
        logging.info(f"Created summary sheet {sheet_name}")

    logging.debug("Processing completed, returning workbook")
    return output_wb