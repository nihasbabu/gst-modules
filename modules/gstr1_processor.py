import json
import os
import datetime
import zipfile
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter

# ----------------------- Global Variables ----------------------- #
# Indian number format (without rupee symbol)
INDIAN_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"

# Section titles
SECTION_TITLES = {
    "B2B,SEZ,DE": "B2B, SEZ, DE Invoices",
    "CDNR": "CDNR - Credit/Debit Notes (Registered)",
    "B2CS": "B2CS - B2C (Others)",
    "NIL": "NIL - Nil Rated, Exempted and Non-GST Supplies",
    "EXP": "EXP - Exports Invoices (with/without payment)",
    "HSN": "HSN - HSN wise details of outward supplies",
    "B2BA": "B2BA - Amended B2B Invoices",
    "CDNUR": "CDNUR - Credit/Debit Notes (Unregistered)",
    "DOC1": "1. Invoices for outward supply",
    "DOC2": "2. Invoices for inward supply from unregistered person",
    "DOC3": "3. Revised Invoice",
    "DOC4": "4. Debit Note",
    "DOC5": "5. Credit Note",
    "DOC6": "6. Receipt voucher",
    "DOC7": "7. Payment Voucher",
    "DOC8": "8. Refund voucher",
    "DOC9": "9. Delivery Challan for job work",
    "DOC10": "10. Delivery Challan for supply on approval",
    "DOC11": "11. Delivery Challan in case of liquid gas",
    "DOC12": "12. Delivery Challan in cases other than by way of supply (excluding at S no. 9 to 11)",
    "AT": "Tax Liability (Advances Received)",
    "TXPD": "Adjustment of Advances",
    "Summary-B2B": "4A-Supplies to registered persons(other than reverse charge)-B2B Regular-Summary",
    "Summary-SEZWP-WOP": "6B-Supplies made to SEZ-SEZWP/SEZWOP Total-Summary",
    "Summary-B2CS": "7-Supplies to unregistered persons-B2CS (Others)-Summary",
    "Summary-CDNR": "9B-Credit/Debit Notes(Registered)-Summary",
    "Summary-NIL": "8-Nil Rated,exempted,non GST supplies-Summary",
    "Summary-EXPWP": "6A–Exports (with payment)-Summary",
    "Summary-EXPWOP": "6A–Exports (without payment)-Summary",
    "Summary-EXP-Total": "6A–Exports (with/without payment)-Summary",
    "Summary-B2BA Total": "9A-Amendment to Supplies made to registered persons in earlier tax period-B2B Amended total-Summary",
    "Summary-CDNUR-B2CL": "9B-Credit/Debit Notes(Unregistered)-B2CL-Summary",
    "Summary-CDNUR-EXPWP": "9B-Credit/Debit Notes(Unregistered)-EXPWP-Summary",
    "Summary-CDNUR-EXPWOP": "9B-Credit/Debit Notes(Unregistered)-EXPWOP-Summary",
    "Summary-CDNUR-TOTAL": "9B-Credit/Debit Notes(Unregistered)-CDNUR-Total-Summary",
    "Summary-HSN": "12-HSN wise outward supplies-Summary",
    "Summary-DOC": "13-Documents issued-Summary",
    "Summary-AT": "11A(1),11A(2)-Advances received-No invoice issued (tax to be added to tax liability)-Summary",
    "Summary-TXPD": "11B(1),11B(2)-Advances received in earlier tax period-Adjusted in this tax period-Summary",
    "B2B,SEZ,DE_sws": "B2B, SEZ, DE Invoices - Sorted Supplier_wise",
    "CDNR_sws": "CDNR - Credit/Debit Notes (Registered) - Sorted Supplier_wise"
}

# Column headers for each section
COLUMN_HEADERS = {
    "B2B,SEZ,DE": [
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice number", "Invoice date",
        "Reporting Month", "Tax type", "Invoice value", "Place of Supply", "Reverse Charge",
        "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN", "Rate", "Taxable Value",
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "IRN", "IRN date", "E-invoice status"
    ],
    "CDNR": [
        "GSTIN/UIN of Recipient", "Receiver Name", "Note Number", "Note Date", "Reporting Month",
        "Note Type", "Place of Supply", "Reverse Charge", "Note Supply Type", "Note Value",
        "Applicable % of Tax Rate", "Rate", "Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess Amount", "IRN", "IRN date", "E-invoice status"
    ],
    "B2CS": [
        "Reporting Month", "Place of Supply", "Rate", "Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess", "Applicable % of Tax Rate",
        "Type", "Supply Type"
    ],
    "NIL": [
        "Reporting Month", "Supply Type", "Nil Rated Supplies", "Exempted(Other than Nil rated/non-GST supply)",
        "Non-GST Supplies"
    ],
    "EXP": [
        "Invoice no", "Invoice date", "Reporting Month", "GST payment", "Supply type",
        "Total Invoice value", "Rate", "Total Taxable Value", "Integrated Tax", "Central Tax",
        "State/UT Tax", "Cess", "IRN", "IRN date"
    ],
    "HSN": [
        "Reporting Month", "HSN/SAC", "No. of Records", "UQC", "Quantity", "Taxable Value",
        "Tax Rate", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"
    ],
    "B2BA": [
        "Recipient GSTIN/UIN", "Revised Invoice no", "Revised Invoice date", "Reporting Month",
        "Revised/Original Invoice no", "Revised/Original Invoice date",
        "Total Invoice value", "Rate", "Total Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess"
    ],
    "CDNUR": [
        "C/D Note No", "C/D Note Date", "Reporting Month", "Note Type", "Type", "Rate",
        "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
        "IRN", "IRN date"
    ],
    "DOC": [
        "Reporting Month", "From (Sr. No.)", "To (Sr. No.)", "Total Number", "Cancelled", "Net Issued"
    ],
    "AT": [
        "Month", "Place of Supply", "Supply Type", "Gross Advance Adjusted",
        "Integrated Tax", "Central Tax", "State/UT Tax", "CESS"
    ],
    "TXPD": [
        "Month", "Place of Supply", "Supply Type", "Gross Advance Adjusted",
        "Integrated Tax", "Central Tax", "State/UT Tax", "CESS"
    ],
    "Summary": [
        "Reporting Month", "No. of Records", "Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess"
    ],
    "Summary-DOC": [
        "Reporting Month", "No. of Records", "Net issued Documents", "Documents issued", "Documents cancelled"
    ]
}

# Column formats
COLUMN_FORMATS = {
    "B2B,SEZ,DE": {
        "Invoice date": "DD-MM-YYYY", "Tax type": "General", "Invoice value": INDIAN_FORMAT,
        "Place of Supply": "#,##0", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "IRN date": "DD-MM-YYYY"
    },
    "CDNR": {
        "Note Date": "DD-MM-YYYY", "Note Value": INDIAN_FORMAT, "Place of Supply": "#,##0",
        "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess Amount": INDIAN_FORMAT,
        "IRN date": "DD-MM-YYYY"
    },
    "B2CS": {
        "Place of Supply": "#,##0", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "NIL": {
        "Nil Rated Supplies": INDIAN_FORMAT, "Exempted(Other than Nil rated/non-GST supply)": INDIAN_FORMAT,
        "Non-GST Supplies": INDIAN_FORMAT
    },
    "EXP": {
        "Invoice no": "General", "Invoice date": "DD-MM-YYYY", "GST payment": "General",
        "Supply type": "General", "Total Invoice value": INDIAN_FORMAT, "Rate": INDIAN_FORMAT, "Total Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "IRN": "General", "IRN date": "DD-MM-YYYY"
    },
    "HSN": {
        "Reporting Month": "General", "HSN/SAC": "#,##0", "No. of Records": "#,##0", "UQC": "General",
        "Quantity": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT, "Tax Rate": "#,##0.00",
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "B2BA": {
        "Recipient GSTIN/UIN": "General", "Revised Invoice no": "General", "Revised Invoice date": "DD-MM-YYYY",
        "Revised/Original Invoice no": "General", "Revised/Original Invoice date": "DD-MM-YYYY",
        "Total Invoice value": INDIAN_FORMAT, "Rate": INDIAN_FORMAT, "Total Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "CDNUR": {
        "C/D Note Date": "DD-MM-YYYY", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess": INDIAN_FORMAT,
        "IRN date": "DD-MM-YYYY"
    },
    "DOC": {
        "Reporting Month": "General", "From (Sr. No.)": "General", "To (Sr. No.)": "General",
        "Total Number": "#,##0", "Cancelled": "#,##0", "Net Issued": "#,##0"
    },
    "AT": {
        "Month": "General", "Place of Supply": "General", "Supply Type": "General",
        "Gross Advance Adjusted": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "CESS": INDIAN_FORMAT
    },
    "TXPD": {
        "Month": "General", "Place of Supply": "General", "Supply Type": "General",
        "Gross Advance Adjusted": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "CESS": INDIAN_FORMAT
    },
    "Summary": {
        "Reporting Month": "General", "No. of Records": "#,##0", "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "Summary-DOC": {
        "Reporting Month": "General", "No. of Records": "#,##0", "Net issued Documents": "#,##0",
        "Documents issued": "#,##0", "Documents cancelled": "#,##0"
    }
}

# Numeric keys for each section
NUMERIC_KEYS_BY_SECTION = {
    "B2B,SEZ,DE": ["Invoice value", "Place of Supply", "Rate", "Taxable Value", "Integrated Tax", "Central Tax",
                   "State/UT Tax", "Cess"],
    "CDNR": ["Note Value", "Place of Supply", "Rate", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax",
             "Cess Amount"],
    "B2CS": ["Place of Supply", "Rate", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "NIL": ["Nil Rated Supplies", "Exempted(Other than Nil rated/non-GST supply)", "Non-GST Supplies"],
    "EXP": ["Total Invoice value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "HSN": ["No. of Records", "Quantity", "Taxable Value", "Tax Rate", "Integrated Tax", "Central Tax", "State/UT Tax",
            "Cess"],
    "B2BA": ["Total Invoice value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "CDNUR": ["Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "DOC": ["Total Number", "Cancelled", "Net Issued"],
    "AT": ["Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax", "CESS"],
    "TXPD": ["Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax", "CESS"],
    "Summary": ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "Summary-DOC": ["No. of Records", "Net issued Documents", "Documents issued", "Documents cancelled"]
}


# ----------------------- Utility Functions ----------------------- #
def parse_filename(filename):
    """Extract month and excluded sections from filename."""
    basename = os.path.basename(filename)
    month_match = re.search(r'GSTR1_(\d{6})', basename)
    month = month_match.group(1) if month_match else None
    excl_match = re.search(r'excluding_([A-Z_]+)', basename)
    excluded = excl_match.group(1).split('_') if excl_match else []
    return month, excluded


def parse_large_filename(filename):
    """Extract month from >500 JSON zip filename."""
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = base.split('_')
    for part in parts:
        if len(part) == 6 and part.isdigit():
            return part
    match = re.search(r'(\d{6})$', base)
    if match:
        return match.group(1)
    return ""


def get_tax_period(ret_str):
    """Convert period string to month name."""
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    ret_str = str(ret_str) if ret_str is not None else ""
    return month_map.get(ret_str[:2], "Unknown") if ret_str and len(ret_str) >= 6 else "Unknown"


def parse_date_string(date_str):
    """Parse date string into datetime.date object."""
    if not date_str:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value, float_2dec=False, int_no_dec=False):
    """Convert value to number with optional formatting."""
    if value is None or value == "":
        return 0
    try:
        num = float(value)
        if int_no_dec:
            return int(num)
        return round(num, 2) if float_2dec else num
    except (ValueError, TypeError):
        return 0


def load_json_data_from_file(file_path, is_zip=False):
    """Load JSON data from file or ZIP."""
    print(f"[DEBUG] Loading JSON from {file_path} (is_zip={is_zip})")
    data_list = []
    try:
        if is_zip:
            with zipfile.ZipFile(file_path, "r") as z:
                json_file = next((name for name in z.namelist() if name.lower().endswith(".json")), None)
                if json_file:
                    with z.open(json_file) as f:
                        data = json.loads(f.read().decode('utf-8'))
                        data["month"] = get_tax_period(data.get("fp", ""))
                        data_list.append(data)
        else:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                period_key = list(data.keys())[0] if data else ""
                data["month"] = get_tax_period(period_key)
                data_list.append(data)
        print(f"[DEBUG] Loaded JSON from {file_path} successfully")
    except Exception as e:
        print(f"[DEBUG] Error loading {file_path}: {e}")
    return data_list


# ----------------------- Extraction Functions (<500 Logic) ----------------------- #
def extract_b2b_entries(data):
    """Extract B2B entries from <500 JSON, emitting a single 'error' row if nested details are missing."""
    print("[DEBUG] Extracting B2B,SEZ,DE section...")
    if not data:
        print("[DEBUG] Extracted B2B,SEZ,DE section...done (no data)")
        return []

    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(
        data.get(period_key, {})
            .get("summary", {})
            .get("data", {})
            .get("ret_period", "")
    )

    suppliers = data.get(period_key, {}).get("sections", {}).get("B2B", {}).get("suppliers", [])
    results = []
    invoice_counts = {}

    for supplier in suppliers:
        sup_info   = supplier.get("supplier", {})
        gstin      = sup_info.get("ctin", "")
        trade_name = sup_info.get("trade_name", "")
        tax_type   = sup_info.get("txp_typ", "")

        for inv in supplier.get("invoiceDetails", []):
            invoice_num   = inv.get("inum", "").strip()
            if not invoice_num:
                continue

            # top level fields
            invoice_date   = parse_date_string(inv.get("idt", ""))
            invoice_value  = parse_number(inv.get("val", ""), float_2dec=True)
            reverse_charge = inv.get("rchrg", "")
            invoice_type   = inv.get("inv_typ", "")
            ecom_gstin     = inv.get("ctin", "")
            irn            = inv.get("irn", "")
            irn_date       = parse_date_string(inv.get("irngendate", ""))
            e_inv_status   = "Yes" if irn else ""

            # amount fields come from these top‑level keys
            top_txval = parse_number(inv.get("invtxval", 0), float_2dec=True)
            top_iamt  = parse_number(inv.get("inviamt", 0), float_2dec=True)
            top_camt  = parse_number(inv.get("invcamt", 0), float_2dec=True)
            top_samt  = parse_number(inv.get("invsamt", 0), float_2dec=True)
            top_csamt = parse_number(inv.get("invcsamt", 0), float_2dec=True)

            nested = inv.get("invoiceDetails", [])

            # if nested missing, emit one row with only Rate="error"
            if not nested or not nested[0].get("inv", []):
                results.append({
                    "GSTIN/UIN of Recipient": gstin,
                    "Receiver Name":           trade_name,
                    "Invoice number":          invoice_num,
                    "Invoice date":            invoice_date,
                    "Reporting Month":         reporting_month,
                    "Tax type":                tax_type,
                    "Invoice value":           invoice_value,
                    "Place of Supply":         gstin[:2] if gstin else "",
                    "Reverse Charge":          reverse_charge,
                    "Applicable % of Tax Rate": None,
                    "Invoice Type":            invoice_type,
                    "E-Commerce GSTIN":        ecom_gstin,
                    "Rate":                    "error",
                    "Taxable Value":           top_txval,
                    "Integrated Tax":          top_iamt,
                    "Central Tax":             top_camt,
                    "State/UT Tax":            top_samt,
                    "Cess":                    top_csamt,
                    "IRN":                     irn,
                    "IRN date":                irn_date,
                    "E-invoice status":        e_inv_status,
                    "highlight":               False
                })
                invoice_counts[invoice_num] = invoice_counts.get(invoice_num, 0) + 1
                continue

            # otherwise, drill into nested itms
            inv_data = nested[0]["inv"][0]
            for item in inv_data.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")):
                    continue

                rate  = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt  = parse_number(itm_det.get("iamt", 0), float_2dec=True)
                camt  = parse_number(itm_det.get("camt", 0), float_2dec=True)
                samt  = parse_number(itm_det.get("samt", 0), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)

                results.append({
                    "GSTIN/UIN of Recipient": gstin,
                    "Receiver Name":           trade_name,
                    "Invoice number":          invoice_num,
                    "Invoice date":            invoice_date,
                    "Reporting Month":         reporting_month,
                    "Tax type":                tax_type,
                    "Invoice value":           invoice_value,
                    "Place of Supply":         gstin[:2] if gstin else "",
                    "Reverse Charge":          reverse_charge,
                    "Applicable % of Tax Rate": None,
                    "Invoice Type":            invoice_type,
                    "E-Commerce GSTIN":        ecom_gstin,
                    "Rate":                    rate,
                    "Taxable Value":           txval,
                    "Integrated Tax":          iamt,
                    "Central Tax":             camt,
                    "State/UT Tax":            samt,
                    "Cess":                    csamt,
                    "IRN":                     irn,
                    "IRN date":                irn_date,
                    "E-invoice status":        e_inv_status,
                    "highlight":               False
                })
                invoice_counts[invoice_num] = invoice_counts.get(invoice_num, 0) + 1

    # highlight any invoice with >1 row
    for row in results:
        if invoice_counts.get(row["Invoice number"], 0) > 1:
            row["highlight"] = True

    print("[DEBUG] Extracted B2B,SEZ,DE section...done")
    return results


def extract_cdnr_entries(data):
    """
    Extract CDNR entries from <500 JSON with nested invoice details,
    emitting a single 'error' row if nested details are missing,
    and pulling amounts from top level when nested is absent.
    """
    print("[DEBUG] Extracting CDNR section...")
    if not data:
        print("[DEBUG] Extracted CDNR section...done (no data)")
        return []

    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(
        data.get(period_key, {})
            .get("summary", {})
            .get("data", {})
            .get("ret_period", "")
    )
    suppliers = data.get(period_key, {}).get("sections", {}).get("CDNR", {}).get("suppliers", [])
    results = []
    note_counts = {}

    for supplier in suppliers:
        s_info     = supplier.get("supplier", {})
        gstin      = s_info.get("ctin", "")
        trade_name = s_info.get("trade_name", "")

        for note in supplier.get("invoiceDetails", []):
            note_num = note.get("nt_num", "").strip()
            if not note_num:
                continue

            # Top-level fields
            note_date   = parse_date_string(note.get("nt_dt", ""))
            note_type   = note.get("ntty", "")
            reverse_charge = note.get("rchrg", "")
            supply_type = note.get("inv_typ", "")
            irn         = note.get("irn", "")
            irn_date    = parse_date_string(note.get("irngendate", ""))
            e_inv_status = "Yes" if irn else ""

            # Pull amounts from top level
            top_txval = parse_number(note.get("invtxval", note.get("val", 0)), float_2dec=True)
            top_iamt  = parse_number(note.get("inviamt", 0), float_2dec=True)
            top_camt  = parse_number(note.get("invcamt", 0), float_2dec=True)
            top_samt  = parse_number(note.get("invsamt", 0), float_2dec=True)
            top_csamt = parse_number(note.get("invcsamt", 0), float_2dec=True)

            nested = note.get("invoiceDetails", [])

            # If nested 'nt' absent, emit one row with Rate="error" and top-level amounts
            if not nested or not nested[0].get("nt", []):
                results.append({
                    "GSTIN/UIN of Recipient": gstin,
                    "Receiver Name":           trade_name,
                    "Note Number":             note_num,
                    "Note Date":               note_date,
                    "Reporting Month":         reporting_month,
                    "Note Type":               note_type,
                    "Place of Supply":         gstin[:2] if gstin else "",
                    "Reverse Charge":          reverse_charge,
                    "Note Supply Type":        supply_type,
                    "Note Value":              parse_number(note.get("val", 0), float_2dec=True),
                    "Applicable % of Tax Rate": None,
                    "Rate":                    "error",
                    "Taxable Value":           top_txval,
                    "Integrated Tax":          top_iamt,
                    "Central Tax":             top_camt,
                    "State/UT Tax":            top_samt,
                    "Cess Amount":             top_csamt,
                    "IRN":                     irn,
                    "IRN date":                irn_date,
                    "E-invoice status":        e_inv_status,
                    "highlight":               False
                })
                note_counts[note_num] = note_counts.get(note_num, 0) + 1
                continue

            # Otherwise, drill into nested 'nt' -> itms
            nt_data = nested[0]["nt"][0]
            for item in nt_data.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")):
                    continue

                rate  = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt  = parse_number(itm_det.get("iamt", 0), float_2dec=True)
                camt  = parse_number(itm_det.get("camt", 0), float_2dec=True)
                samt  = parse_number(itm_det.get("samt", 0), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)

                results.append({
                    "GSTIN/UIN of Recipient": gstin,
                    "Receiver Name":           trade_name,
                    "Note Number":             note_num,
                    "Note Date":               note_date,
                    "Reporting Month":         reporting_month,
                    "Note Type":               note_type,
                    "Place of Supply":         gstin[:2] if gstin else "",
                    "Reverse Charge":          reverse_charge,
                    "Note Supply Type":        supply_type,
                    "Note Value":              parse_number(note.get("val", 0), float_2dec=True),
                    "Applicable % of Tax Rate": None,
                    "Rate":                    rate,
                    "Taxable Value":           txval,
                    "Integrated Tax":          iamt,
                    "Central Tax":             camt,
                    "State/UT Tax":            samt,
                    "Cess Amount":             csamt,
                    "IRN":                     irn,
                    "IRN date":                irn_date,
                    "E-invoice status":        e_inv_status,
                    "highlight":               False
                })
                note_counts[note_num] = note_counts.get(note_num, 0) + 1

    # Highlight duplicates
    for row in results:
        if note_counts.get(row["Note Number"], 0) > 1:
            row["highlight"] = True

    print("[DEBUG] Extracted CDNR section...done")
    return results


def extract_b2cs_entries(data):
    """Extract B2CS entries from <500 JSON."""
    print("[DEBUG] Extracting B2CS section...")
    if not data:
        print("[DEBUG] Extracted B2CS section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
    inner = data.get(period_key, {})
    b2cs_obj = inner.get("sections", {}).get("B2CS", {})
    items = b2cs_obj.get("invoiceDetails", []) if isinstance(b2cs_obj, dict) else b2cs_obj if isinstance(b2cs_obj,
                                                                                                         list) else []
    results = []
    for item in items:
        row = {
            "Reporting Month": reporting_month,
            "Place of Supply": parse_number(item.get("pos", ""), int_no_dec=True),
            "Rate": parse_number(item.get("rt", ""), float_2dec=True),
            "Taxable Value": parse_number(item.get("invtxval", ""), float_2dec=True),
            "Integrated Tax": parse_number(item.get("inviamt", ""), float_2dec=True),
            "Central Tax": parse_number(item.get("invcamt", ""), float_2dec=True),
            "State/UT Tax": parse_number(item.get("invsamt", ""), float_2dec=True),
            "Cess": parse_number(item.get("invcsamt", ""), float_2dec=True),
            "Applicable % of Tax Rate": None,
            "Type": item.get("typ", ""),
            "Supply Type": item.get("sply_ty", "")
        }
        results.append(row)
    print("[DEBUG] Extracted B2CS section...done")
    return results


def extract_nil_entries(data):
    """Extract NIL entries from <500 JSON."""
    print("[DEBUG] Extracting NIL section...")
    if not data:
        print("[DEBUG] Extracted NIL section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
    inner = data.get(period_key, {})
    nil_obj = inner.get("sections", {}).get("NIL", {})
    items = nil_obj.get("inv", nil_obj.get("invoiceDetails", [])) if isinstance(nil_obj,
                                                                                dict) else nil_obj if isinstance(
        nil_obj, list) else []
    results = []
    for inv in items:
        row = {
            "Reporting Month": reporting_month,
            "Supply Type": inv.get("sply_ty", ""),
            "Nil Rated Supplies": parse_number(inv.get("nil_amt", ""), float_2dec=True),
            "Exempted(Other than Nil rated/non-GST supply)": parse_number(inv.get("expt_amt", ""), float_2dec=True),
            "Non-GST Supplies": parse_number(inv.get("ngsup_amt", ""), float_2dec=True)
        }
        results.append(row)
    print("[DEBUG] Extracted NIL section...done")
    return results


def extract_exp_entries(data):
    """
    Extract EXP entries from <500 JSON, where everything lives under a period key (e.g. "032025").
    Emits a single 'error' row if nested details are missing (Rate only), pulling amounts from top level.
    Highlights any invoice with multiple rows.
    """
    from collections import Counter

    entries = []

    # 1) Find the period key (skip any top‑level helper like "month")
    period_key = next((k for k in data.keys() if k != "month"), None)
    if not period_key:
        return []

    # 2) Read reporting month
    summary = data[period_key].get("summary", {}).get("data", {})
    reporting_month = get_tax_period(summary.get("ret_period", ""))

    # 3) Pull out the EXP section
    exp_section = data[period_key].get("sections", {}).get("EXP", {})

    # 4) Build entries
    for invoice in exp_section.get("invoiceDetails", []):
        inum         = invoice.get("inum", "")
        idt          = invoice.get("idt", "")
        val          = invoice.get("val", 0.00)
        irn          = invoice.get("irn", "")
        irn_date     = invoice.get("irngendate", "")
        gst_payment  = invoice.get("exp_typ", "")
        supply_type  = invoice.get("srctyp", "")

        # top-level taxable & tax amounts
        top_txval = parse_number(invoice.get("invtxval", invoice.get("val", 0)), float_2dec=True)
        top_iamt  = parse_number(invoice.get("inviamt", 0), float_2dec=True)
        top_camt  = parse_number(invoice.get("invcamt", 0), float_2dec=True)
        top_samt  = parse_number(invoice.get("invsamt", 0), float_2dec=True)
        top_csamt = parse_number(invoice.get("invcsamt", 0), float_2dec=True)

        nested_list = invoice.get("invoiceDetails", [])
        inv_array   = nested_list[0].get("inv", []) if nested_list else []

        # if nested 'inv' missing or empty → emit single error row
        if not inv_array:
            entries.append({
                "Invoice no":           inum,
                "Invoice date":         idt,
                "Reporting Month":      reporting_month,
                "GST payment":          gst_payment,
                "Supply type":          supply_type,
                "Total Invoice value":  val,
                "Rate":                 "error",
                "Total Taxable Value":  top_txval,
                "Integrated Tax":       top_iamt,
                "Central Tax":          top_camt,
                "State/UT Tax":         top_samt,
                "Cess":                 top_csamt,
                "IRN":                  irn,
                "IRN date":             irn_date,
                "highlight":            False
            })
            continue

        # otherwise, drill into nested inv → itms
        for inv in inv_array:
            for item in inv.get("itms", []):
                # only emit rows with both rt and txval
                if not all(k in item for k in ("rt", "txval")):
                    continue

                rt    = item["rt"]
                txval = item["txval"]
                iamt  = item.get("iamt", 0.00)
                camt  = item.get("camt", 0.00)
                samt  = item.get("samt", 0.00)
                csamt = item.get("csamt", 0.00)

                entries.append({
                    "Invoice no":           inum,
                    "Invoice date":         idt,
                    "Reporting Month":      reporting_month,
                    "GST payment":          gst_payment,
                    "Supply type":          supply_type,
                    "Total Invoice value":  val,
                    "Rate":                 rt,
                    "Total Taxable Value":  txval,
                    "Integrated Tax":       iamt,
                    "Central Tax":          camt,
                    "State/UT Tax":         samt,
                    "Cess":                 csamt,
                    "IRN":                  irn,
                    "IRN date":             irn_date,
                    "highlight":            False
                })

    # 5) Highlight any invoice that appears more than once
    counts = Counter(row["Invoice no"] for row in entries)
    for row in entries:
        if counts.get(row["Invoice no"], 0) > 1:
            row["highlight"] = True

    return entries


def extract_hsn_entries(data_list):
    """Extract and aggregate HSN entries from all <500 JSON files."""
    print("[DEBUG] Extracting HSN section...")
    if not data_list:
        print("[DEBUG] Extracted HSN section...done (empty data)")
        return []
    hsn_dict = {}
    for data in data_list:
        if not data or not isinstance(data, dict):
            continue
        period_key = list(data.keys())[0]
        reporting_month = get_tax_period(
            data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
        hsn_obj = data.get(period_key, {}).get("sections", {}).get("HSN", {})
        items = hsn_obj.get("invoiceDetails", []) if isinstance(hsn_obj, dict) else hsn_obj if isinstance(hsn_obj,
                                                                                                          list) else []
        for item in items:
            hsn_sac = item.get("hsn_sc", "").strip()
            uqc = item.get("uqc", "").strip()
            if not hsn_sac or not uqc:
                continue
            key = (reporting_month, hsn_sac, uqc)
            if key not in hsn_dict:
                hsn_dict[key] = {
                    "Reporting Month": reporting_month,
                    "HSN/SAC": hsn_sac,
                    "No. of Records": 0,
                    "UQC": uqc,
                    "Quantity": 0,
                    "Taxable Value": 0,
                    "Tax Rate": parse_number(item.get("rt", ""), float_2dec=True),
                    "Integrated Tax": 0,
                    "Central Tax": 0,
                    "State/UT Tax": 0,
                    "Cess": 0
                }
            hsn_dict[key]["No. of Records"] += 1
            hsn_dict[key]["Quantity"] += parse_number(item.get("qty", ""), float_2dec=True)
            hsn_dict[key]["Taxable Value"] += parse_number(item.get("txval", ""), float_2dec=True)
            hsn_dict[key]["Integrated Tax"] += parse_number(item.get("iamt", ""), float_2dec=True)
            hsn_dict[key]["Central Tax"] += parse_number(item.get("camt", ""), float_2dec=True)
            hsn_dict[key]["State/UT Tax"] += parse_number(item.get("samt", ""), float_2dec=True)
            hsn_dict[key]["Cess"] += parse_number(item.get("csamt", ""), float_2dec=True)
    results = list(hsn_dict.values())
    financial_order = ["April", "May", "June", "July", "August", "September", "October", "November", "December",
                       "January", "February", "March"]
    results.sort(key=lambda x: (
    financial_order.index(x["Reporting Month"]) if x["Reporting Month"] in financial_order else 999, x["HSN/SAC"]))
    print("[DEBUG] Extracted HSN section...done")
    return results


def extract_b2ba_entries(data):
    """
    Extract B2BA entries from <500 JSON where everything is nested under a period key
    (e.g. "062024"). Emits a single 'error' row if nested details are missing (Rate only),
    pulling amounts from top level, and highlights any invoice with multiple rows.
    """
    from collections import Counter

    entries = []
    invoice_counts = {}

    # 1) locate the period key
    period_key = next((k for k in data.keys() if k != "month"), None)
    if not period_key:
        return []

    section = data[period_key] \
                .get("sections", {}) \
                .get("B2BA", {})

    # 2) get the reporting month
    ret_period = data[period_key] \
                   .get("summary", {}) \
                   .get("data", {}) \
                   .get("ret_period", "")
    reporting_month = get_tax_period(ret_period)

    for inv_wrap in section.get("invoiceDetails", []):
        # top‑level fields
        inum  = inv_wrap.get("inum", "")
        idt   = inv_wrap.get("idt", "")
        oinum = inv_wrap.get("oinum", "")
        oidt  = inv_wrap.get("oidt", "")
        val   = parse_number(inv_wrap.get("val", 0), float_2dec=True)
        ctin  = inv_wrap.get("ctin", "")

        # pull top‑level taxable & tax amounts
        top_txval = parse_number(inv_wrap.get("invtxval", inv_wrap.get("val", 0)), float_2dec=True)
        top_iamt  = parse_number(inv_wrap.get("inviamt", 0), float_2dec=True)
        top_camt  = parse_number(inv_wrap.get("invcamt", 0), float_2dec=True)
        top_samt  = parse_number(inv_wrap.get("invsamt", 0), float_2dec=True)
        top_csamt = parse_number(inv_wrap.get("invcsamt", 0), float_2dec=True)

        # prepare highlight count
        invoice_counts[oinum] = invoice_counts.get(oinum, 0) + 0  # will increment per-row below

        # 3) nested details → inv → itms
        nested = inv_wrap.get("invoiceDetails", [])
        inv_list = nested[0].get("inv", []) if nested else []

        # if nested missing or empty → emit single error row
        if not inv_list:
            entries.append({
                "Recipient GSTIN/UIN":           ctin,
                "Revised Invoice no":            inum,
                "Revised Invoice date":          idt,
                "Reporting Month":               reporting_month,
                "Revised/Original Invoice no":   oinum,
                "Revised/Original Invoice date": oidt,
                "Total Invoice value":           val,
                "Rate":                          "error",
                "Total Taxable Value":           top_txval,
                "Integrated Tax":                top_iamt,
                "Central Tax":                   top_camt,
                "State/UT Tax":                  top_samt,
                "Cess":                          top_csamt,
                "highlight":                     False
            })
            invoice_counts[oinum] += 1
            continue

        # otherwise, process each itm_det that has both keys
        for inv in inv_list:
            for item in inv.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")):
                    continue

                rt    = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt  = parse_number(itm_det.get("iamt", 0), float_2dec=True)
                camt  = parse_number(itm_det.get("camt", 0), float_2dec=True)
                samt  = parse_number(itm_det.get("samt", 0), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)

                entries.append({
                    "Recipient GSTIN/UIN":           ctin,
                    "Revised Invoice no":            inum,
                    "Revised Invoice date":          idt,
                    "Reporting Month":               reporting_month,
                    "Revised/Original Invoice no":   oinum,
                    "Revised/Original Invoice date": oidt,
                    "Total Invoice value":           val,
                    "Rate":                          rt,
                    "Total Taxable Value":           txval,
                    "Integrated Tax":                iamt,
                    "Central Tax":                   camt,
                    "State/UT Tax":                  samt,
                    "Cess":                          csamt,
                    "highlight":                     False
                })
                invoice_counts[oinum] += 1

    # 5) highlight any invoice that appears more than once
    counts = Counter(invoice_counts)
    for row in entries:
        if counts.get(row["Revised/Original Invoice no"], 0) > 1:
            row["highlight"] = True

    return entries


def extract_cdnur_entries(data):
    """
    Extract CDNUR entries from <500 JSON where everything is nested under a period key
    (e.g. "052024"). Emits a single 'error' row if nested details are missing (Rate only),
    pulling amounts from top level, and highlights any note with multiple rows.
    """
    from collections import Counter

    entries = []
    note_counts = {}

    # 1) locate the period key
    period_key = next((k for k in data.keys() if k != "month"), None)
    if not period_key:
        return []

    # 2) get reporting month
    summary_data = data[period_key].get("summary", {}).get("data", {})
    reporting_month = get_tax_period(summary_data.get("ret_period", ""))

    # 3) pull out the CDNUR section
    cdnur_section = data[period_key].get("sections", {}).get("CDNUR", {})

    # 4) iterate each top‑level note
    for note in cdnur_section.get("invoiceDetails", []):
        nt_num    = note.get("nt_num", "")
        nt_dt     = note.get("nt_dt", "")
        ntty      = note.get("ntty", "")
        typ       = note.get("typ", "")
        irn       = note.get("irn", "")
        irn_date  = note.get("irngendate", "")

        # top-level amounts
        top_txval = parse_number(note.get("invtxval", note.get("val", 0)), float_2dec=True)
        top_iamt  = parse_number(note.get("inviamt", 0), float_2dec=True)
        top_camt  = parse_number(note.get("invcamt", 0), float_2dec=True)
        top_samt  = parse_number(note.get("invsamt", 0), float_2dec=True)
        top_csamt = parse_number(note.get("invcsamt", 0), float_2dec=True)

        # prepare nested itms list
        nested = note.get("invoiceDetails", [])
        itms = nested[0].get("itms", []) if nested else []

        # if missing nested details → emit single error row
        if not itms:
            entries.append({
                "C/D Note No":        nt_num,
                "C/D Note Date":      nt_dt,
                "Reporting Month":    reporting_month,
                "Note Type":          ntty,
                "Type":               typ,
                "Rate":               "error",
                "Taxable Value":      top_txval,
                "Integrated Tax":     top_iamt,
                "Central Tax":        top_camt,
                "State/UT Tax":       top_samt,
                "Cess":               top_csamt,
                "IRN":                irn,
                "IRN date":           irn_date,
                "highlight":          False
            })
            note_counts[nt_num] = note_counts.get(nt_num, 0) + 1
            continue

        # otherwise, process each itm_det that has both rt and txval
        for item in itms:
            itm_det = item.get("itm_det", {})
            if not all(k in itm_det for k in ("rt", "txval")):
                continue

            rt    = itm_det["rt"]
            txval = parse_number(itm_det["txval"], float_2dec=True)
            iamt  = parse_number(itm_det.get("iamt", 0), float_2dec=True)
            camt  = parse_number(itm_det.get("camt", 0), float_2dec=True)
            samt  = parse_number(itm_det.get("samt", 0), float_2dec=True)
            csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)

            entries.append({
                "C/D Note No":        nt_num,
                "C/D Note Date":      nt_dt,
                "Reporting Month":    reporting_month,
                "Note Type":          ntty,
                "Type":               typ,
                "Rate":               rt,
                "Taxable Value":      txval,
                "Integrated Tax":     iamt,
                "Central Tax":        camt,
                "State/UT Tax":       samt,
                "Cess":               csamt,
                "IRN":                irn,
                "IRN date":           irn_date,
                "highlight":          False
            })
            note_counts[nt_num] = note_counts.get(nt_num, 0) + 1

    # 5) highlight any note number that appears more than once
    counts = Counter(row["C/D Note No"] for row in entries)
    for row in entries:
        if counts.get(row["C/D Note No"], 0) > 1:
            row["highlight"] = True

    return entries


def extract_doc_entries(data):
    """Extract DOC entries from <500 JSON."""
    print("[DEBUG] Extracting DOC section...")
    if not data:
        print("[DEBUG] Extracted DOC section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
    doc_obj = data.get(period_key, {}).get("sections", {}).get("DOC", {})
    items = doc_obj.get("invoiceDetails", []) if isinstance(doc_obj, dict) else doc_obj if isinstance(doc_obj,
                                                                                                      list) else []
    results = []
    for item in items:
        doc_type = item.get("doc_typ", "")
        for doc in item.get("docs", []):
            row = {
                "Reporting Month": reporting_month,
                "From (Sr. No.)": doc.get("from", ""),
                "To (Sr. No.)": doc.get("to", ""),
                "Total Number": parse_number(doc.get("totnum", ""), int_no_dec=True),
                "Cancelled": parse_number(doc.get("cancel", ""), int_no_dec=True),
                "Net Issued": parse_number(doc.get("net_issue", ""), int_no_dec=True),
                "doc_type": doc_type
            }
            results.append(row)
    print("[DEBUG] Extracted DOC section...done")
    return results


def extract_at_entries(data):
    """Extract AT entries from <500 JSON."""
    print("[DEBUG] Extracting AT section...")
    if not data:
        print("[DEBUG] Extracted AT section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
    at_obj = data.get(period_key, {}).get("sections", {}).get("AT", {})
    items = at_obj.get("invoiceDetails", []) if isinstance(at_obj, dict) else at_obj if isinstance(at_obj, list) else []
    results = []
    for item in items:
        row = {
            "Month": reporting_month,
            "Place of Supply": item.get("pos", ""),
            "Supply Type": item.get("sply_ty", ""),
            "Gross Advance Adjusted": parse_number(item.get("invadamt", ""), float_2dec=True),
            "Integrated Tax": parse_number(item.get("inviamt", ""), float_2dec=True),
            "Central Tax": parse_number(item.get("invcamt", ""), float_2dec=True),
            "State/UT Tax": parse_number(item.get("invsamt", ""), float_2dec=True),
            "CESS": parse_number(item.get("invcsamt", ""), float_2dec=True)
        }
        results.append(row)
    print("[DEBUG] Extracted AT section...done")
    return results


def extract_txpd_entries(data):
    """Extract TXPD entries from <500 JSON."""
    print("[DEBUG] Extracting TXPD section...")
    if not data:
        print("[DEBUG] Extracted TXPD section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = get_tax_period(data.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period", ""))
    txpd_obj = data.get(period_key, {}).get("sections", {}).get("TXPD", {})
    items = txpd_obj.get("invoiceDetails", []) if isinstance(txpd_obj, dict) else txpd_obj if isinstance(txpd_obj,
                                                                                                         list) else []
    results = []
    for item in items:
        row = {
            "Month": reporting_month,
            "Place of Supply": item.get("pos", ""),
            "Supply Type": item.get("sply_ty", ""),
            "Gross Advance Adjusted": parse_number(item.get("invadamt", ""), float_2dec=True),
            "Integrated Tax": parse_number(item.get("inviamt", ""), float_2dec=True),
            "Central Tax": parse_number(item.get("invcamt", ""), float_2dec=True),
            "State/UT Tax": parse_number(item.get("invsamt", ""), float_2dec=True),
            "CESS": parse_number(item.get("invcsamt", ""), float_2dec=True)
        }
        results.append(row)
    print("[DEBUG] Extracted TXPD section...done")
    return results


# ----------------------- Extraction Functions (>500 Logic) ----------------------- #
def extract_b2b_entries_large(data):
    """Extract B2B entries from >500 JSON with multi-rate highlighting."""
    print("[DEBUG] Extracting B2B,SEZ,DE section (large JSON)...")
    if not data or not isinstance(data, dict):
        print("[DEBUG] Extracted B2B,SEZ,DE section (large JSON)...done (empty data)")
        return []
    supplier_state = data.get('gstin', '')[:2]
    reporting_month = get_tax_period(data.get("fp", ""))
    valid_rates = [0.00, 5.00, 12.00, 18.00, 28.00]
    results = []
    for buyer in data.get("b2b", []):
        gstin = buyer.get("ctin", "")
        for inv in buyer.get("inv", []):
            invoice_num = inv.get("inum", "").strip()
            if not invoice_num:
                continue
            rates = {parse_number(item["itm_det"].get("rt", ""), float_2dec=True)
                     for item in inv.get("itms", []) if parse_number(item["itm_det"].get("rt", "")) > 0}
            multi_rate_highlight = len(rates) >= 2
            inv_typ = inv.get("inv_typ", "")
            invoice_base = {
                "GSTIN/UIN of Recipient": gstin,
                "Receiver Name": "",
                "Invoice number": invoice_num,
                "Invoice date": parse_date_string(inv.get("idt", "")),
                "Reporting Month": reporting_month,
                "Tax type": inv_typ,
                "Invoice value": parse_number(inv.get("val", ""), float_2dec=True),
                "Place of Supply": inv.get("pos", ""),
                "Reverse Charge": inv.get("rchrg", ""),
                "Applicable % of Tax Rate": None,
                "Invoice Type": inv_typ,
                "E-Commerce GSTIN": "",
                "IRN": inv.get("irn", ""),
                "IRN date": parse_date_string(inv.get("irngendate", "")),
                "E-invoice status": "Yes" if inv.get("irn") else "",
                "highlight": False
            }
            for item in inv.get("itms", []):
                itm_det = item.get("itm_det", {})
                rate = parse_number(itm_det.get("rt", ""), float_2dec=True)
                for valid in valid_rates:
                    if abs(rate - valid) <= 0.02:
                        rate = valid
                        break
                row = invoice_base.copy()
                row.update({
                    "Rate": rate,
                    "Taxable Value": parse_number(itm_det.get("txval", ""), float_2dec=True),
                    "Integrated Tax": parse_number(itm_det.get("iamt", ""), float_2dec=True) if inv.get(
                        "pos") != supplier_state else 0,
                    "Central Tax": parse_number(itm_det.get("camt", ""), float_2dec=True) if inv.get(
                        "pos") == supplier_state else 0,
                    "State/UT Tax": parse_number(itm_det.get("samt", ""), float_2dec=True) if inv.get(
                        "pos") == supplier_state else 0,
                    "Cess": parse_number(itm_det.get("csamt", ""), float_2dec=True),
                    "highlight": multi_rate_highlight or (rate not in valid_rates)
                })
                results.append(row)
    print("[DEBUG] Extracted B2B,SEZ,DE section (large JSON)...done")
    return results


# ----------------------- Summary Calculation Functions ----------------------- #
def safe_add(current_value, new_value):
    """Safely add only if new_value is numeric."""
    if isinstance(new_value, (int, float)):
        return current_value + new_value
    return current_value


def calculate_monthly_summary(
    data,
    date_key,
    taxable_key,
    iamt_key,
    camt_key,
    samt_key,
    cess_key,
    invoice_key=None,
    processed_months=None
):
    """
    Calculate monthly summary for a section, skipping any non-numeric inputs.
    Optionally counts unique invoices if invoice_key is provided.
    """
    summary = {}

    # Initialize summary for all processed months
    if processed_months:
        for month in processed_months:
            summary[month] = {
                "record_count": 0,
                "taxable_value": 0.0,
                "integrated_tax": 0.0,
                "central_tax": 0.0,
                "state_ut_tax": 0.0,
                "cess": 0.0,
                "unique_invoices": set() if invoice_key else None
            }

    # Aggregate data
    for row in data:
        month = row.get(date_key)
        if not isinstance(month, str):
            continue
        if month not in summary:
            summary[month] = {
                "record_count": 0,
                "taxable_value": 0.0,
                "integrated_tax": 0.0,
                "central_tax": 0.0,
                "state_ut_tax": 0.0,
                "cess": 0.0,
                "unique_invoices": set() if invoice_key else None
            }

        # Count rows or unique invoices
        if invoice_key and row.get(invoice_key):
            summary[month]["unique_invoices"].add(row[invoice_key])
        else:
            summary[month]["record_count"] += 1

        # Safely add each numeric field
        summary[month]["taxable_value"] = safe_add(summary[month]["taxable_value"], row.get(taxable_key))
        summary[month]["integrated_tax"] = safe_add(summary[month]["integrated_tax"], row.get(iamt_key))
        summary[month]["central_tax"]    = safe_add(summary[month]["central_tax"], row.get(camt_key))
        summary[month]["state_ut_tax"]   = safe_add(summary[month]["state_ut_tax"], row.get(samt_key))
        summary[month]["cess"]           = safe_add(summary[month]["cess"], row.get(cess_key))

    # Build sorted result by financial year
    financial_order = [
        "April","May","June","July","August","September",
        "October","November","December","January","February","March"
    ]
    result = []
    for m in financial_order:
        if m in summary:
            cnt = (
                len(summary[m]["unique_invoices"])
                if summary[m]["unique_invoices"] is not None
                else summary[m]["record_count"]
            )
            result.append({
                "Reporting Month": m,
                "No. of Records": cnt,
                "Taxable Value": summary[m]["taxable_value"],
                "Integrated Tax": summary[m]["integrated_tax"],
                "Central Tax": summary[m]["central_tax"],
                "State/UT Tax": summary[m]["state_ut_tax"],
                "Cess": summary[m]["cess"]
            })
    return result


def make_cdnr_values_negative(summary):
    """Negate CDNR values."""
    for row in summary:
        for key in ["Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"]:
            if row.get(key, 0) != 0:
                row[key] = -row[key]
    return summary


def make_cdnur_values_negative(summary):
    """Negate CDNUR values."""
    for row in summary:
        for key in ["Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"]:
            if row.get(key, 0) != 0:
                row[key] = -row[key]
    return summary


# ----------------------- Excel Report Generation ----------------------- #
def create_excel_report(data_dict, wb, ignore_warnings=False):
    """Create Excel sheets only for sections with data, and mark tabs red if any 'error' cells appear."""
    print("[DEBUG] Creating detailed sheets...")
    for section, rows in data_dict.items():
        if section == "DOC":
            continue
        # skip sections without data (or all zeros) unless ignoring warnings
        if not rows or (not ignore_warnings and not any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in rows
            for key in NUMERIC_KEYS_BY_SECTION.get(section, [])
        )):
            continue

        sheet_name = f"R1-{section}"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        # write header row
        cols = COLUMN_HEADERS.get(section, [])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(row=1, column=1, value=SECTION_TITLES.get(section, section)).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for idx, col_name in enumerate(cols, start=1):
            hdr = ws.cell(row=2, column=idx, value=col_name)
            hdr.font = Font(bold=True)
            hdr.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "B3"

        sheet_has_error = False

        # write data rows
        for r_idx, row_data in enumerate(rows, start=3):
            is_highlight = row_data.get("highlight", False)
            for c_idx, col_name in enumerate(cols, start=1):
                val = row_data.get(col_name, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)

                # if we see an "error", flag the sheet and style it bold red
                if val == "error":
                    sheet_has_error = True
                    cell.font = Font(color="FF0000", bold=True)
                # otherwise, if highlighted, style bold red
                elif is_highlight:
                    cell.font = Font(color="FF0000", bold=True)

                # apply number format only to numeric values
                if (
                    section in COLUMN_FORMATS
                    and col_name in COLUMN_FORMATS[section]
                    and isinstance(val, (int, float))
                ):
                    cell.number_format = COLUMN_FORMATS[section][col_name]

        # autofit columns
        for idx, col_name in enumerate(cols, start=1):
            col_letter = get_column_letter(idx)
            max_len = len(col_name)
            for row in range(3, ws.max_row + 1):
                v = ws.cell(row=row, column=idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = max(15, max_len + 1)

        # if any errors, paint the tab red
        if sheet_has_error:
            ws.sheet_properties.tabColor = "FF0000"

        print(f"[DEBUG] Created sheet {sheet_name}")
    print("[DEBUG] Finished creating detailed sheets")


def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    """Create or replace a worksheet with headers."""
    print(f"[DEBUG] Creating/replacing sheet {sheet_name}...")
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=title_text).font = Font(bold=True, size=12)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=idx, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3"
    print(f"[DEBUG] Created sheet {sheet_name}")
    return ws


def fill_worksheet_data(ws, columns, data, start_row=3):
    """Fill worksheet with data."""
    red_font = Font(color="FF0000", bold=True)
    numeric_cols = {"No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
                    "Net issued Documents", "Documents issued", "Documents cancelled", "Total Number", "Cancelled",
                    "Net Issued"}
    for row_idx, row_dict in enumerate(data, start_row):
        is_highlight = row_dict.get("highlight", False)
        for col_idx, col_name in enumerate(columns, 1):
            val = row_dict.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = float(val) if col_name in numeric_cols and val else val
            if is_highlight:
                cell.font = red_font


def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    """Apply formatting and adjust column widths."""
    for col_idx, col_name in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        if col_format_map and col_name in col_format_map:
            for row in range(start_row, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).number_format = col_format_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 1)


# ----------------------- Main Processing Function ----------------------- #
def process_gstr1(small_files, large_files, excluded_sections_by_month, template_path, save_path,
                  ignore_warnings=False):
    """Process GSTR1 data and generate Excel report."""
    print("[DEBUG] Starting GSTR1 processing...")
    combined_data = {
        key: [] for key in [
            "B2B,SEZ,DE", "CDNR", "B2CS", "NIL", "EXP", "HSN", "B2BA", "CDNUR", "DOC", "AT", "TXPD"
        ]
    }
    all_data_list = []
    processed_months = set()

    # Process <500 JSON files
    print("[DEBUG] Processing small JSON files...")
    for file in small_files:
        month, excluded = parse_filename(file)
        data_list = load_json_data_from_file(file)
        all_data_list.extend(data_list)
        for data in data_list:
            reporting_month = data.get("month", "Unknown")
            if reporting_month != "Unknown":
                processed_months.add(reporting_month)
            if "B2B" not in excluded:
                combined_data["B2B,SEZ,DE"].extend(extract_b2b_entries(data))
            if "CDNR" not in excluded:
                combined_data["CDNR"].extend(extract_cdnr_entries(data))
            if "B2CS" not in excluded:
                combined_data["B2CS"].extend(extract_b2cs_entries(data))
            if "NIL" not in excluded:
                combined_data["NIL"].extend(extract_nil_entries(data))
            if "EXP" not in excluded:
                combined_data["EXP"].extend(extract_exp_entries(data))
            if "B2BA" not in excluded:
                combined_data["B2BA"].extend(extract_b2ba_entries(data))
            if "CDNUR" not in excluded:
                combined_data["CDNUR"].extend(extract_cdnur_entries(data))
            if "DOC" not in excluded:
                combined_data["DOC"].extend(extract_doc_entries(data))
            if "AT" not in excluded:
                combined_data["AT"].extend(extract_at_entries(data))
            if "TXPD" not in excluded:
                combined_data["TXPD"].extend(extract_txpd_entries(data))
            if "HSN" not in excluded:
                print("[DEBUG] Collecting HSN data for aggregation...")
                # No direct extraction here; data collected in all_data_list
                print("[DEBUG] Collected HSN data for aggregation...done")
    print("[DEBUG] Finished processing small JSON files")

    # Process >500 JSON files
    print("[DEBUG] Processing large JSON files...")
    for month, (filepath, _) in large_files.items():
        if not filepath:
            continue
        excluded = excluded_sections_by_month.get(month, [])
        if "B2B" not in excluded:
            for data in load_json_data_from_file(filepath, is_zip=True):
                reporting_month = data.get("month", "Unknown")
                if reporting_month != "Unknown":
                    processed_months.add(reporting_month)
                combined_data["B2B,SEZ,DE"].extend(extract_b2b_entries_large(data))
    print("[DEBUG] Finished processing large JSON files")

    # Aggregate HSN data
    print("[DEBUG] Aggregating HSN data...")
    combined_data["HSN"] = extract_hsn_entries(all_data_list)
    print("[DEBUG] HSN data aggregation completed")

    # Check for data
    has_data = any(combined_data[section] for section in combined_data if section != "DOC")
    if not has_data and not ignore_warnings:
        raise ValueError("No data found in provided JSON files.")
    print("[DEBUG] Data validation completed")

    # Sort data
    print("[DEBUG] Sorting data...")
    financial_order = [
        "April", "May", "June", "July", "August", "September",
        "October", "November", "December", "January", "February", "March"
    ]

    # B2B: month → invoice date
    combined_data["B2B,SEZ,DE"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("Invoice date") or datetime.datetime.max
        )
    )

    # CDNR: month → note date
    combined_data["CDNR"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("Note Date") or datetime.datetime.max
        )
    )

    # B2CS & NIL already month‑only
    combined_data["B2CS"].sort(
        key=lambda x: financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                                   "") in financial_order else 999
    )
    combined_data["NIL"].sort(
        key=lambda x: financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                                   "") in financial_order else 999
    )

    # EXP: month → invoice date
    combined_data["EXP"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("Invoice date") or datetime.datetime.max
        )
    )

    # HSN: month → HSN/SAC
    combined_data["HSN"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("HSN/SAC", "")
        )
    )

    # B2BA: month → revised invoice date
    combined_data["B2BA"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("Revised Invoice date") or datetime.datetime.max
        )
    )

    # CDNUR: month → C/D note date
    combined_data["CDNUR"].sort(
        key=lambda x: (
            financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                         "") in financial_order else 999,
            x.get("C/D Note Date") or datetime.datetime.max
        )
    )

    # DOC, AT & TXPD are month‑only
    combined_data["DOC"].sort(
        key=lambda x: financial_order.index(x.get("Reporting Month", "")) if x.get("Reporting Month",
                                                                                   "") in financial_order else 999
    )
    combined_data["AT"].sort(
        key=lambda x: financial_order.index(x.get("Month", "")) if x.get("Month", "") in financial_order else 999
    )
    combined_data["TXPD"].sort(
        key=lambda x: financial_order.index(x.get("Month", "")) if x.get("Month", "") in financial_order else 999
    )

    print("[DEBUG] Data sorting completed")

    # Create workbook
    print("[DEBUG] Initializing workbook...")
    wb = load_workbook(template_path) if template_path and os.path.exists(template_path) else Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    print("[DEBUG] Workbook initialized")

    # Generate detailed sheets
    create_excel_report(combined_data, wb, ignore_warnings)

    # Generate DOC detailed sheets
    print("[DEBUG] Generating document-specific sheets...")
    doc_data = combined_data.get("DOC", [])
    doc_types = {
        "Invoices for outward supply": "DOC1",
        "Invoices for inward supply from unregistered person": "DOC2",
        "Revised Invoice": "DOC3",
        "Debit Note": "DOC4",
        "Credit Note": "DOC5",
        "Receipt voucher": "DOC6",
        "Payment Voucher": "DOC7",
        "Refund voucher": "DOC8",
        "Delivery Challan for job work": "DOC9",
        "Delivery Challan for supply on approval": "DOC10",
        "Delivery Challan in case of liquid gas": "DOC11",
        "Delivery Challan in cases other than by way of supply (excluding at S no. 9 to 11)": "DOC12"
    }
    for doc_type, sheet_suffix in doc_types.items():
        filtered_rows = [row for row in doc_data if row["doc_type"] == doc_type]
        if not filtered_rows or (not ignore_warnings and not any(
                isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
                for row in filtered_rows for key in NUMERIC_KEYS_BY_SECTION.get("DOC", []))):
            continue
        sheet_name = f"R1-{sheet_suffix}"
        ws = wb.create_sheet(sheet_name)
        cols = COLUMN_HEADERS["DOC"]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(row=1, column=1, value=SECTION_TITLES[sheet_suffix]).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "B3"
        for row_idx, row_data in enumerate(filtered_rows, 3):
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
                if "DOC" in COLUMN_FORMATS and col_name in COLUMN_FORMATS["DOC"]:
                    cell.number_format = COLUMN_FORMATS["DOC"][col_name]
        for col_idx, col_name in enumerate(cols, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for row in range(3, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max(15, max_len + 1)
        print(f"[DEBUG] Created sheet {sheet_name}")
    print("[DEBUG] Finished generating document-specific sheets")

    # Generate supplier-wise sorted sheets (R1-CDNR_sws and R1-B2B,SEZ,DE_sws)
    print("[DEBUG] Generating supplier-wise sorted sheets...")
    for section in ["CDNR", "B2B,SEZ,DE"]:
        rows = combined_data.get(section, [])
        if not rows or (not ignore_warnings and not any(
                isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
                for row in rows for key in NUMERIC_KEYS_BY_SECTION.get(section, []))):
            continue
        sorted_rows = sorted(
            rows,
            key=lambda x: (
                x.get("Receiver Name", "") if x.get("Receiver Name", "") else x.get("GSTIN/UIN of Recipient", ""),
                x.get("GSTIN/UIN of Recipient", "") if not x.get("Receiver Name", "") else ""
            )
        )
        sheet_name = f"R1-{section}_sws"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        cols = COLUMN_HEADERS.get(section, [])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(row=1, column=1, value=SECTION_TITLES.get(f"{section}_sws", section)).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "B3"
        for row_idx, row_data in enumerate(sorted_rows, 3):
            is_highlight = row_data.get("highlight", False)
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
                if is_highlight:
                    cell.font = Font(color="FF0000", bold=True)
                if section in COLUMN_FORMATS and col_name in COLUMN_FORMATS[section]:
                    cell.number_format = COLUMN_FORMATS[section][col_name]
        for col_idx, col_name in enumerate(cols, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for row in range(3, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    max_len = max(max_len, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max(15, max_len + 1)
        print(f"[DEBUG] Created sheet {sheet_name}")
    print("[DEBUG] Finished generating supplier-wise sorted sheets")

    # Generate Summary Sheets
    print("[DEBUG] Generating summary sheets...")
    summary_columns = COLUMN_HEADERS["Summary"]
    summary_ws_list = []

    # B2B Summary (NT and CO or R)
    b2b_nt_co_r_data = [row for row in combined_data.get("B2B,SEZ,DE", [])
                        if row.get("Tax type") in ["NT", "CO", "R"]]
    b2b_summary = calculate_monthly_summary(
        b2b_nt_co_r_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        invoice_key="Invoice number",
        processed_months=processed_months
    )
    if b2b_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in b2b_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-B2B",
            SECTION_TITLES["Summary-B2B"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, b2b_summary)
        summary_ws_list.append(ws)

    # SEZ Summary (SEZ, SEZWOP, SEZWP)
    sez_data = [row for row in combined_data.get("B2B,SEZ,DE", [])
                if row.get("Tax type") in ["SEZ", "SEZWOP", "SEZWP", "SEWP", "SEWOP"]]
    sez_summary = calculate_monthly_summary(
        sez_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        processed_months=processed_months
    )
    if sez_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in sez_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-SEZWP-WOP",
            SECTION_TITLES["Summary-SEZWP-WOP"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, sez_summary)
        summary_ws_list.append(ws)

    # B2CS Summary
    b2cs_summary = calculate_monthly_summary(
        combined_data.get("B2CS", []),
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        processed_months=processed_months
    )
    if b2cs_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in b2cs_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-B2CS",
            SECTION_TITLES["Summary-B2CS"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, b2cs_summary)
        summary_ws_list.append(ws)

    # CDNR Summary
    cdnr_summary = calculate_monthly_summary(
        combined_data.get("CDNR", []),
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess Amount",
        invoice_key="Note Number",
        processed_months=processed_months
    )
    cdnr_summary = make_cdnr_values_negative(cdnr_summary)
    if cdnr_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in cdnr_summary for key in
            ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-CDNR",
            SECTION_TITLES["Summary-CDNR"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, cdnr_summary)
        summary_ws_list.append(ws)

    # NIL Summary
    nil_summary = calculate_monthly_summary(
        combined_data.get("NIL", []),
        "Reporting Month",
        "Nil Rated Supplies",
        None,
        None,
        None,
        None,
        processed_months=processed_months
    )
    if nil_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in nil_summary for key in ["No. of Records", "Taxable Value"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-NIL",
            SECTION_TITLES["Summary-NIL"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, nil_summary)
        summary_ws_list.append(ws)

    # AT Summary
    at_summary = calculate_monthly_summary(
        combined_data.get("AT", []),
        "Month",
        "Gross Advance Adjusted",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "CESS",
        processed_months=processed_months
    )
    if at_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in at_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-AT",
            SECTION_TITLES["Summary-AT"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, at_summary)
        summary_ws_list.append(ws)

    # TXPD Summary
    txpd_summary = calculate_monthly_summary(
        combined_data.get("TXPD", []),
        "Month",
        "Gross Advance Adjusted",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "CESS",
        processed_months=processed_months
    )
    if txpd_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in txpd_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-TXPD",
            SECTION_TITLES["Summary-TXPD"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, txpd_summary)
        summary_ws_list.append(ws)

    # HSN Summary
    hsn_data = combined_data.get("HSN", [])
    hsn_summary = calculate_monthly_summary(
        hsn_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        processed_months=processed_months
    )
    if hsn_summary:
        for summary_row in hsn_summary:
            month = summary_row["Reporting Month"]
            summary_row["No. of Records"] = sum(
                row["No. of Records"] for row in hsn_data if row["Reporting Month"] == month)
    if hsn_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in hsn_summary for key in ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-HSN",
            SECTION_TITLES["Summary-HSN"],
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, hsn_summary)
        summary_ws_list.append(ws)

    # DOC Summary
    doc_summary_data = []
    if doc_data or processed_months:
        month_dict = {month: {"doc_types": set(), "net_issued": 0, "total_num": 0, "cancelled": 0}
                      for month in processed_months}
        for row in doc_data:
            month = row["Reporting Month"]
            if month not in month_dict:
                month_dict[month] = {"doc_types": set(), "net_issued": 0, "total_num": 0, "cancelled": 0}
            month_dict[month]["doc_types"].add(row["doc_type"])
            month_dict[month]["net_issued"] += row["Net Issued"]
            month_dict[month]["total_num"] += row["Total Number"]
            month_dict[month]["cancelled"] += row["Cancelled"]
        doc_summary_data = [
            {
                "Reporting Month": month,
                "No. of Records": len(values["doc_types"]),
                "Net issued Documents": values["net_issued"],
                "Documents issued": values["total_num"],
                "Documents cancelled": values["cancelled"]
            }
            for month, values in month_dict.items()
        ]
        doc_summary_data.sort(
            key=lambda x: financial_order.index(x["Reporting Month"]) if x[
                                                                             "Reporting Month"] in financial_order else 999
        )
        if doc_summary_data and (ignore_warnings or any(
                isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
                for row in doc_summary_data for key in ["No. of Records", "Net issued Documents", "Documents issued", "Documents cancelled"])):
            ws = create_or_replace_sheet(
                wb,
                "R1-Summary-DOC",
                SECTION_TITLES["Summary-DOC"],
                COLUMN_HEADERS["Summary-DOC"]
            )
            fill_worksheet_data(ws, COLUMN_HEADERS["Summary-DOC"], doc_summary_data)
            apply_format_and_autofit(ws, COLUMN_HEADERS["Summary-DOC"], col_format_map=COLUMN_FORMATS["Summary-DOC"])
            summary_ws_list.append(ws)

    # B2BA Summary
    b2ba_summary = calculate_monthly_summary(
        combined_data.get("B2BA", []),
        "Reporting Month",
        "Total Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        invoice_key="Revised/Original Invoice no",
        processed_months=processed_months
    )
    if b2ba_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in b2ba_summary for key in
            ["No. of Records", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-B2BA Total",
            SECTION_TITLES.get("Summary-B2BA Total", "B2BA Summary Total"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, b2ba_summary)
        summary_ws_list.append(ws)

    # EXP Summaries
    # Summary-EXPWP (GST Payment = WPAY)
    expwp_data = [row for row in combined_data.get("EXP", []) if row.get("GST payment") == "WPAY"]
    expwp_summary = calculate_monthly_summary(
        expwp_data,
        "Reporting Month",
        "Total Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        invoice_key="Invoice no",
        processed_months=processed_months
    )
    if expwp_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in expwp_summary for key in
            ["No. of Records", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-EXPWP",
            SECTION_TITLES.get("Summary-EXPWP", "EXPWP Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, expwp_summary)
        summary_ws_list.append(ws)

    # Summary-EXPWOP (GST Payment = WOPAY)
    expwop_data = [row for row in combined_data.get("EXP", []) if row.get("GST payment") == "WOPAY"]
    expwop_summary = calculate_monthly_summary(
        expwop_data,
        "Reporting Month",
        "Total Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        invoice_key="Invoice no",
        processed_months=processed_months
    )
    if expwop_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in expwop_summary for key in
            ["No. of Records", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-EXPWOP",
            SECTION_TITLES.get("Summary-EXPWOP", "EXPWOP Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, expwop_summary)
        summary_ws_list.append(ws)

    # Summary-EXP-Total
    exp_total_summary = calculate_monthly_summary(
        combined_data.get("EXP", []),
        "Reporting Month",
        "Total Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess",
        invoice_key="Invoice no",
        processed_months=processed_months
    )
    if exp_total_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in exp_total_summary for key in
            ["No. of Records", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-EXP-Total",
            SECTION_TITLES.get("Summary-EXP-Total", "EXP Total Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, exp_total_summary)
        summary_ws_list.append(ws)

    # CDNUR Summaries
    # Summary-CDNUR-B2CL (Type = B2CL)
    cdnur_b2cl_data = [row for row in combined_data.get("CDNUR", []) if row.get("Type") == "B2CL"]
    cdnur_b2cl_summary = calculate_monthly_summary(
        cdnur_b2cl_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess Amount",
        invoice_key="C/D Note No",
        processed_months=processed_months
    )
    cdnur_b2cl_summary = make_cdnr_values_negative(cdnur_b2cl_summary)
    if cdnur_b2cl_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in cdnur_b2cl_summary for key in
            ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-CDNUR-B2CL",
            SECTION_TITLES.get("Summary-CDNUR-B2CL", "CDNUR B2CL Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, cdnur_b2cl_summary)
        summary_ws_list.append(ws)

    # Summary-CDNUR-EXPWP (Type = EXPWP)
    cdnur_expwp_data = [row for row in combined_data.get("CDNUR", []) if row.get("Type") == "EXPWP"]
    cdnur_expwp_summary = calculate_monthly_summary(
        cdnur_expwp_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess Amount",
        invoice_key="C/D Note No",
        processed_months=processed_months
    )
    cdnur_expwp_summary = make_cdnr_values_negative(cdnur_expwp_summary)
    if cdnur_expwp_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in cdnur_expwp_summary for key in
            ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-CDNUR-EXPWP",
            SECTION_TITLES.get("Summary-CDNUR-EXPWP", "CDNUR EXPWP Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, cdnur_expwp_summary)
        summary_ws_list.append(ws)

    # Summary-CDNUR-EXPWOP (Type = EXPWOP)
    cdnur_expwop_data = [row for row in combined_data.get("CDNUR", []) if row.get("Type") == "EXPWOP"]
    cdnur_expwop_summary = calculate_monthly_summary(
        cdnur_expwop_data,
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess Amount",
        invoice_key="C/D Note No",
        processed_months=processed_months
    )
    cdnur_expwop_summary = make_cdnr_values_negative(cdnur_expwop_summary)
    if cdnur_expwop_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in cdnur_expwop_summary for key in
            ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-CDNUR-EXPWOP",
            SECTION_TITLES.get("Summary-CDNUR-EXPWOP", "CDNUR EXPWOP Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, cdnur_expwop_summary)
        summary_ws_list.append(ws)

    # Summary-CDNUR-TOTAL
    cdnur_total_summary = calculate_monthly_summary(
        combined_data.get("CDNUR", []),
        "Reporting Month",
        "Taxable Value",
        "Integrated Tax",
        "Central Tax",
        "State/UT Tax",
        "Cess Amount",
        invoice_key="C/D Note No",
        processed_months=processed_months
    )
    cdnur_total_summary = make_cdnr_values_negative(cdnur_total_summary)
    if cdnur_total_summary and (ignore_warnings or any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in cdnur_total_summary for key in
            ["No. of Records", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"])):
        ws = create_or_replace_sheet(
            wb,
            "R1-Summary-CDNUR-TOTAL",
            SECTION_TITLES.get("Summary-CDNUR-TOTAL", "CDNUR Total Summary"),
            summary_columns
        )
        fill_worksheet_data(ws, summary_columns, cdnur_total_summary)
        summary_ws_list.append(ws)

    # Apply formatting to summary sheets
    print("[DEBUG] Applying formatting to summary sheets...")
    for ws in summary_ws_list:
        columns = ws["A2":f"{get_column_letter(ws.max_column)}2"][0]
        column_names = [cell.value for cell in columns]
        apply_format_and_autofit(ws, column_names, col_format_map=COLUMN_FORMATS.get("Summary", {}))
    print("[DEBUG] Finished formatting summary sheets")

    # Save workbook
    print(f"[DEBUG] Saving workbook to {save_path}...")
    wb.save(save_path)
    print(f"[DEBUG] Workbook saved successfully to {save_path}")

    print("[DEBUG] GSTR1 processing completed")
    return wb